"""
modules/billing/services/product_import_service.py
---------------------------------------------------
Bulk Product & Service Import Service — Phase 5B

Supports:
  - CSV and XLSX file parsing
  - Column auto-detection and user-defined mapping
  - Validation of all product fields (with org-scoped category/tax lookups)
  - Duplicate detection scoped to current organization
  - Preview with short-lived TTL cache (30 min, powered by cachetools)
  - Transactional confirm with partial-success result reporting
  - Auto-create missing categories (configurable)
  - Template generation (CSV + XLSX) with accepted-values annotations
  - Full audit logging for every import operation
  - No cross-tenant leakage: all DB queries scoped to organization_id
"""

from __future__ import annotations

import csv
import io
import logging
import re
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from cachetools import TTLCache
from sqlalchemy.orm import Session

from app.modules.billing.models import BillingAuditAction, Product, ProductCategory
from app.modules.billing.repositories.catalog import (
    ProductCategoryRepository,
    ProductRepository,
)
from app.modules.billing.services.audit_service import BillingAuditService
from app.modules.billing.services.product_service import (
    PRODUCT_ALLOWED_FIELDS,
    ProductService,
    _resolve_org_currency,
)
from app.modules.billing.utils.currency_utils import VALID_CURRENCY_CODES as VALID_CURRENCIES

logger = logging.getLogger("zoiko")

# ---------------------------------------------------------------------------
# Supported file columns → product field mapping
# ---------------------------------------------------------------------------

FIELD_ALIASES: Dict[str, str] = {
    # name
    "name": "name",
    "product name": "name",
    "service name": "name",
    "item name": "name",
    "title": "name",
    # code / sku
    "code": "code",
    "sku": "code",
    "sku / code": "code",
    "product code": "code",
    "item code": "code",
    "reference": "code",
    # type
    "type": "product_type",
    "product type": "product_type",
    "item type": "product_type",
    "service type": "product_type",
    # category
    "category": "category",
    "category name": "category",
    "product category": "category",
    # description
    "description": "description",
    "details": "description",
    "notes": "description",
    # unit
    "unit": "unit_label",
    "unit label": "unit_label",
    "unit / meter": "unit_label",
    "uom": "unit_label",
    # currency
    "currency": "currency",
    "currency code": "currency",
    # unit price / default price
    "unit price": "default_price",
    "default price": "default_price",
    "price": "default_price",
    "selling price": "default_price",
    "rate": "default_price",
    # tax category
    "tax category": "tax_category",
    "tax rate": "tax_percentage",
    "tax %": "tax_percentage",
    "vat rate": "tax_percentage",
    # default discount
    "default discount": "default_discount",
    "discount": "default_discount",
    "discount %": "default_discount",
    "default discount %": "default_discount",
    # country
    "country": "country",
    "country code": "country",
    # gst/vat group
    "gst/vat group": "gst_vat_group",
    "gst vat group": "gst_vat_group",
    "tax group": "gst_vat_group",
    "vat group": "gst_vat_group",
    # invoice description
    "invoice description": "invoice_description",
    "invoice note": "invoice_description",
    # active status
    "status": "status",
    "active": "status",
    "is active": "status",
    "enabled": "status",
    # brand
    "brand": "brand",
    # original price
    "original price": "original_price",
    "list price": "original_price",
    "mrp": "original_price",
    # billing frequency
    "billing frequency": "billing_frequency",
    "frequency": "billing_frequency",
}

# Enterprise-scale safety limits. Enforced before/while parsing so a huge or
# maliciously crafted file (e.g. an XLSX whose small compressed size expands to
# an enormous number of rows) can never force the whole file to be read into
# memory or processed row-by-row unbounded.
MAX_IMPORT_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 20_000

VALID_PRODUCT_TYPES = {"service", "good", "subscription", "usage", "retainer", "other"}
VALID_BILLING_FREQUENCIES = {
    "one_time", "monthly", "quarterly", "yearly", "usage_based", "recurring",
}
VALID_STATUSES = {"active", "inactive"}
# VALID_CURRENCIES was previously a separately-maintained hardcoded set here
# that had drifted from the canonical CurrencyCode enum (e.g. it accepted
# currencies like TRY/RUB/ILS that the rest of the app — payments,
# validation — rejects, while rejecting NPR, which the rest of the app
# accepts). Now imported as an alias of the same canonical set every other
# billing service validates against.

# ---------------------------------------------------------------------------
# Global TTL Cache — keyed by (session_id, organization_id)
# Expires after 30 minutes automatically
# ---------------------------------------------------------------------------

_PREVIEW_CACHE: TTLCache = TTLCache(maxsize=512, ttl=1800)  # 30-minute TTL


# ---------------------------------------------------------------------------
# Result data classes (plain dicts; Pydantic schemas are in schemas.py)
# ---------------------------------------------------------------------------

def _row_result(
    row_index: int,
    raw: Dict[str, Any],
    status: str,  # "valid" | "duplicate" | "invalid" | "warning"
    errors: Optional[List[str]] = None,
    warnings: Optional[List[str]] = None,
    mapped: Optional[Dict[str, Any]] = None,
    matched_id: Optional[int] = None,
    matched_code: Optional[str] = None,
) -> Dict[str, Any]:
    return {
        "row_index": row_index,
        "raw_data": raw,
        "mapped_data": mapped or {},
        "status": status,
        "errors": errors or [],
        "warnings": warnings or [],
        "matched_existing_id": matched_id,
        "matched_existing_code": matched_code,
    }


# ---------------------------------------------------------------------------
# File parsing helpers
# ---------------------------------------------------------------------------

def _parse_csv(file_bytes: bytes, max_rows: Optional[int] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse CSV bytes → (headers, rows). Stops reading (rather than truncating
    silently) once max_rows is exceeded, so a huge file is never fully
    materialized in memory."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows: List[Dict[str, str]] = []
    for row in reader:
        if max_rows is not None and len(rows) >= max_rows:
            raise ValueError(
                f"This file has more than {max_rows:,} data rows. "
                f"Please split it into smaller files and import them separately."
            )
        rows.append(dict(row))
    return list(headers), rows


def _parse_xlsx(file_bytes: bytes, max_rows: Optional[int] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse XLSX bytes → (headers, rows). Uses openpyxl's read_only (streaming)
    mode and stops iterating once max_rows is exceeded — protects against a
    small-on-disk XLSX that expands to an enormous number of rows/cells."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX import") from exc

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            return [], []
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
        rows: List[Dict[str, str]] = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue  # skip blank rows
            if max_rows is not None and len(rows) >= max_rows:
                raise ValueError(
                    f"This file has more than {max_rows:,} data rows. "
                    f"Please split it into smaller files and import them separately."
                )
            row_dict = {}
            for h, v in zip(headers, raw_row):
                row_dict[h] = str(v).strip() if v is not None else ""
            rows.append(row_dict)
        return headers, rows
    finally:
        wb.close()


def _check_file_size(file_bytes: bytes) -> None:
    size = len(file_bytes)
    if size > MAX_IMPORT_FILE_SIZE_BYTES:
        raise ValueError(
            f"File is too large ({size / (1024 * 1024):.1f} MB). "
            f"The maximum allowed size is {MAX_IMPORT_FILE_SIZE_BYTES // (1024 * 1024)} MB — "
            f"please split it into smaller files and import them separately."
        )


def _auto_map_columns(headers: List[str]) -> Dict[str, str]:
    """Auto-detect file column → product field mappings."""
    mapping: Dict[str, str] = {}
    for h in headers:
        if not h:
            continue
        key = h.lower().strip()
        canonical = FIELD_ALIASES.get(key) or FIELD_ALIASES.get(key.rstrip(" *").strip())
        if canonical:
            mapping[h] = canonical
    return mapping


# ---------------------------------------------------------------------------
# Value normalisation helpers
# ---------------------------------------------------------------------------

def _normalize_status(val: str) -> Optional[bool]:
    """Convert status string → is_active bool."""
    v = val.lower().strip()
    if v in {"true", "yes", "1", "active", "enabled"}:
        return True
    if v in {"false", "no", "0", "inactive", "disabled"}:
        return False
    return None  # unparseable


def _normalize_decimal(val: str) -> Optional[float]:
    try:
        return float(val.replace(",", "").strip())
    except (ValueError, AttributeError):
        return None


# ---------------------------------------------------------------------------
# Main service
# ---------------------------------------------------------------------------

class ProductImportService:
    """
    Handles the full lifecycle of a bulk product import:
      parse → preview → (cache) → confirm → audit
    """

    def __init__(self, db: Session):
        self.db = db
        self.repo = ProductRepository(db)
        self.cat_repo = ProductCategoryRepository(db)
        self.audit = BillingAuditService(db)

    # ------------------------------------------------------------------
    # STEP 1: Parse file and return raw columns + sample rows
    # ------------------------------------------------------------------

    def parse_file(
        self,
        file_bytes: bytes,
        filename: str,
    ) -> Dict[str, Any]:
        """
        Parse a CSV or XLSX file.
        Returns:
          detected_columns: list of file column headers
          suggested_mapping: auto-detected column → field mapping
          sample_rows: first 5 rows for display
          total_data_rows: row count (excluding header)
        """
        _check_file_size(file_bytes)
        fname_lower = filename.lower()
        if fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls"):
            headers, rows = _parse_xlsx(file_bytes, max_rows=MAX_IMPORT_ROWS)
        elif fname_lower.endswith(".csv"):
            headers, rows = _parse_csv(file_bytes, max_rows=MAX_IMPORT_ROWS)
        else:
            raise ValueError(f"Unsupported file format. Please upload a .csv or .xlsx file.")

        suggested_mapping = _auto_map_columns(headers)

        return {
            "detected_columns": headers,
            "suggested_mapping": suggested_mapping,
            "sample_rows": rows[:5],
            "total_data_rows": len(rows),
        }

    # ------------------------------------------------------------------
    # STEP 2: Validate + preview — returns session token + preview result
    # ------------------------------------------------------------------

    def preview_import(
        self,
        file_bytes: bytes,
        filename: str,
        column_map: Dict[str, str],
        organization_id: int,
        duplicate_strategy: str = "skip",   # skip | overwrite | create_copy | review
        auto_create_categories: bool = True,
    ) -> Dict[str, Any]:
        """
        Validate all rows, detect duplicates, build preview summary.
        Caches the parsed+validated rows under a session_id (30 min TTL).

        Returns:
          session_id: opaque token for confirm step
          expires_at: ISO timestamp when session expires
          total: total rows
          valid: rows that will import cleanly
          duplicate: rows that match existing SKU/name
          invalid: rows with hard validation errors
          warning: rows with soft warnings
          rows: per-row detail
          summary_stats: dict of counts
        """
        _check_file_size(file_bytes)
        fname_lower = filename.lower()
        if fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls"):
            headers, raw_rows = _parse_xlsx(file_bytes, max_rows=MAX_IMPORT_ROWS)
        elif fname_lower.endswith(".csv"):
            headers, raw_rows = _parse_csv(file_bytes, max_rows=MAX_IMPORT_ROWS)
        else:
            raise ValueError("Unsupported file format")

        # Auto-detect common/template columns, then let explicit user overrides win.
        effective_map = dict(_auto_map_columns(headers))
        effective_map.update(column_map or {})

        org_currency = _resolve_org_currency(self.db, organization_id)

        # Pre-fetch existing categories for the org (for lookup + auto-create)
        existing_categories = self._get_org_category_map(organization_id)

        mapped_rows = []
        for i, raw_row in enumerate(raw_rows):
            row_num = i + 1
            mapped_rows.append((
                row_num,
                raw_row,
                *self._map_and_validate_row(
                raw=raw_row,
                column_map=effective_map,
                row_index=row_num,
                organization_id=organization_id,
                org_currency=org_currency,
                existing_categories=existing_categories,
                auto_create_categories=auto_create_categories,
                ),
            ))

        duplicate_candidates = self.repo.list_matching_codes_or_names(
            organization_id=organization_id,
            codes=[mapped.get("code") for _, _, mapped, errors, _ in mapped_rows if not errors and mapped.get("code")],
            names=[mapped.get("name") for _, _, mapped, errors, _ in mapped_rows if not errors and mapped.get("name")],
        )
        duplicate_by_code = {product.code: product for product in duplicate_candidates if product.code}
        duplicate_by_name = {product.name: product for product in duplicate_candidates if product.name}

        processed: List[Dict[str, Any]] = []
        counts = {"valid": 0, "duplicate": 0, "invalid": 0, "warning": 0}

        for row_num, raw_row, mapped, errors, warnings in mapped_rows:

            if errors:
                processed.append(_row_result(row_num, raw_row, "invalid", errors=errors, warnings=warnings, mapped=mapped))
                counts["invalid"] += 1
                continue

            # Duplicate detection (SKU)
            code = mapped.get("code")
            name = mapped.get("name")
            existing = duplicate_by_code.get(code) if code else None
            existing = existing or (duplicate_by_name.get(name) if name else None)
            dup_status = existing is not None
            matched_id = existing.id if existing else None
            matched_code = existing.code if existing else None

            if dup_status:
                processed.append(_row_result(
                    row_num, raw_row, "duplicate",
                    warnings=[f"Duplicate SKU '{code}'" if code else f"Duplicate name '{name}'"] + warnings,
                    mapped=mapped,
                    matched_id=matched_id,
                    matched_code=matched_code,
                ))
                counts["duplicate"] += 1
            elif warnings:
                processed.append(_row_result(row_num, raw_row, "warning", warnings=warnings, mapped=mapped))
                counts["warning"] += 1
            else:
                processed.append(_row_result(row_num, raw_row, "valid", mapped=mapped))
                counts["valid"] += 1

        session_id = str(uuid.uuid4())
        expires_at = datetime.utcnow().isoformat() + "Z+1800s"

        # Cache: store (rows, duplicate_strategy, auto_create_categories, organization_id)
        _PREVIEW_CACHE[(session_id, organization_id)] = {
            "rows": processed,
            "raw_rows": raw_rows,
            "column_map": effective_map,
            "duplicate_strategy": duplicate_strategy,
            "auto_create_categories": auto_create_categories,
            "organization_id": organization_id,
        }

        return {
            "session_id": session_id,
            "expires_at": expires_at,
            "total": len(raw_rows),
            "valid": counts["valid"],
            "duplicate": counts["duplicate"],
            "invalid": counts["invalid"],
            "warning": counts["warning"],
            "rows": processed,
            "summary_stats": counts,
        }

    # ------------------------------------------------------------------
    # STEP 3: Confirm import (transactional, partial-success model)
    # ------------------------------------------------------------------

    def confirm_import(
        self,
        session_id: str,
        organization_id: int,
        user_id: int,
        duplicate_strategy: str = "skip",       # global: skip | overwrite | create_copy
        per_row_actions: Optional[Dict[int, str]] = None,  # {row_index: action} for review mode
        offset: int = 0,
        batch_size: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Commit the import.

        Partial-success model:
          - Each row is processed independently in its own savepoint.
          - Failures accumulate in 'failed' list but do not abort the batch.
          - The caller sees a detailed per-row result.
          - Audit events are logged for the batch start + end.

        Batching (offset/batch_size): each underlying Product create/update is
        its own committed transaction (ProductService/ProductRepository — not
        modified here), so a single confirm_import call over many thousands of
        rows means many thousands of sequential commits in one HTTP request,
        risking a request timeout with no feedback in between. Passing
        batch_size lets a caller process the cached row set in slices across
        multiple requests instead — each call only processes
        rows[offset:offset+batch_size], the session cache is only evicted once
        the final slice completes, and the response reports whether more
        batches remain so the caller can show real progress. Omitting
        batch_size (the default) processes every row in one call, identical to
        this method's original behavior.

        Returns: ImportSummaryResult dict (plus total_rows/next_offset/is_complete).
        """
        cache_key = (session_id, organization_id)
        cached = _PREVIEW_CACHE.get(cache_key)
        if not cached:
            raise ValueError(
                "Import session has expired or is invalid. "
                "Please re-upload your file and preview again."
            )

        if cached["organization_id"] != organization_id:
            raise PermissionError("Session does not belong to this organization.")

        all_rows: List[Dict[str, Any]] = cached["rows"]
        total_rows = len(all_rows)
        rows = all_rows[offset:offset + batch_size] if batch_size is not None else all_rows[offset:]
        batch_end = offset + len(rows)
        is_complete = batch_end >= total_rows

        auto_create = cached.get("auto_create_categories", True)
        per_row_actions = per_row_actions or {}
        product_svc = ProductService(self.db)

        # Audit: import started (only once, on the first batch)
        if offset == 0:
            self.audit.log(
                organization_id, user_id,
                BillingAuditAction.CREATE, "CatalogImport", None,
                new_values={"session_id": session_id, "total_rows": total_rows, "strategy": duplicate_strategy},
            )

        imported: List[int] = []
        skipped: List[int] = []
        failed: List[Dict[str, Any]] = []
        warning_rows: List[int] = []

        for row in rows:
            row_idx = row["row_index"]
            status = row["status"]
            mapped = row.get("mapped_data", {})
            matched_id = row.get("matched_existing_id")

            # Determine effective action for this row
            if status == "invalid":
                failed.append({"row": row_idx, "error": "; ".join(row.get("errors", ["Validation failed"]))})
                continue

            if status == "warning":
                warning_rows.append(row_idx)

            if status == "duplicate":
                # Resolve action: per-row override > global strategy
                action = per_row_actions.get(row_idx, duplicate_strategy)
                if action == "skip":
                    skipped.append(row_idx)
                    continue
                elif action == "overwrite" and matched_id:
                    try:
                        self._update_product_from_mapped(
                            product_svc, matched_id, organization_id, user_id, mapped, auto_create,
                        )
                        imported.append(row_idx)
                    except Exception as exc:
                        failed.append({"row": row_idx, "error": str(exc)})
                    continue
                elif action == "create_copy":
                    mapped = self._make_unique_code_and_name(mapped, organization_id)
                    # falls through to create below
                else:
                    skipped.append(row_idx)
                    continue

            # Create new product
            try:
                self._create_product_from_mapped(
                    product_svc, organization_id, user_id, mapped, auto_create,
                )
                imported.append(row_idx)
            except Exception as exc:
                failed.append({"row": row_idx, "error": str(exc)})

        # Only evict the cache once the final batch has been processed — an
        # in-progress multi-batch import still needs the remaining rows.
        if is_complete:
            _PREVIEW_CACHE.pop(cache_key, None)

        # Audit: batch completed (the final batch's log line effectively
        # summarizes that batch, not the whole import — each batch is its own
        # auditable unit of work, consistent with the partial-success model).
        self.audit.log(
            organization_id, user_id,
            BillingAuditAction.UPDATE, "CatalogImport", None,
            new_values={
                "session_id": session_id,
                "batch_offset": offset,
                "batch_rows": len(rows),
                "is_complete": is_complete,
                "imported": len(imported),
                "skipped": len(skipped),
                "failed": len(failed),
                "warnings": len(warning_rows),
            },
        )

        return {
            "imported": len(imported),
            "skipped": len(skipped),
            "failed": len(failed),
            "warnings": len(warning_rows),
            "imported_row_indices": imported,
            "skipped_row_indices": skipped,
            "failed_details": failed,
            "warning_row_indices": warning_rows,
            "total_rows": total_rows,
            "next_offset": None if is_complete else batch_end,
            "is_complete": is_complete,
        }

    # ------------------------------------------------------------------
    # TEMPLATE GENERATION
    # ------------------------------------------------------------------

    def generate_template(self, fmt: str) -> Tuple[bytes, str]:
        """
        Generate a downloadable import template.
        Returns (file_bytes, mimetype).
        """
        headers = [
            "Name *", "SKU / Code *", "Type *", "Category",
            "Description", "Unit", "Currency", "Unit Price",
            "Default Discount %", "Country", "GST/VAT Group",
            "Invoice Description", "Status",
            "Billing Frequency", "Brand",
        ]
        example_rows = [
            [
                "Website Development",
                "SVC-WEB-001",
                "service",
                "Digital Services",
                "Full-stack web development service",
                "hours",
                "USD",
                "150.00",
                "10",
                "US",
                "STANDARD",
                "Professional web development billed hourly",
                "active",
                "one_time",
                "Zoiko",
            ],
            [
                "Monthly SaaS License",
                "SUB-SAAS-001",
                "subscription",
                "Software",
                "SaaS platform monthly subscription",
                "licenses",
                "USD",
                "99.00",
                "0",
                "",
                "",
                "Monthly SaaS license fee",
                "active",
                "monthly",
                "",
            ],
        ]
        notes = [
            [
                "Required. Product or service name.",
                "Required. Unique code/SKU per org.",
                "Required. Values: service | good | subscription | usage | retainer | other",
                "Optional. Auto-created if missing.",
                "Optional.",
                "Optional. e.g. hours, seats, licenses",
                "Optional. 3-letter code e.g. USD, EUR, INR. Defaults to org currency.",
                "Optional. Numeric. e.g. 99.00",
                "Optional. 0–100.",
                "Optional. 2-letter ISO code e.g. US, IN, GB.",
                "Optional. e.g. STANDARD, REDUCED, EXEMPT.",
                "Optional. Default text shown on invoices.",
                "Optional. active | inactive. Defaults to active.",
                "Optional. one_time | monthly | quarterly | yearly | usage_based | recurring",
                "Optional.",
            ]
        ]

        if fmt == "xlsx":
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # --- Data sheet ---
            ws = wb.active
            ws.title = "Products"

            header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            req_fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(h) + 4)

            for row_data in example_rows:
                ws.append(row_data)
                for col_idx in range(1, len(row_data) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).fill = req_fill

            # --- Notes sheet ---
            ws_notes = wb.create_sheet("Field Notes")
            ws_notes.append(["Field", "Notes / Accepted Values"])
            ws_notes["A1"].font = Font(bold=True)
            ws_notes["B1"].font = Font(bold=True)
            for h, note in zip(headers, notes[0]):
                ws_notes.append([h, note])
            ws_notes.column_dimensions["A"].width = 25
            ws_notes.column_dimensions["B"].width = 80

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else:  # csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(example_rows)
            output.write("\n# Field Notes:\n")
            for h, note in zip(headers, notes[0]):
                output.write(f"# {h}: {note}\n")
            return output.getvalue().encode("utf-8"), "text/csv"

    # ------------------------------------------------------------------
    # EXPORT
    # ------------------------------------------------------------------

    def export_catalog(
        self,
        organization_id: int,
        fmt: str = "csv",
        filters: Optional[Dict[str, Any]] = None,
        ids: Optional[List[int]] = None,
    ) -> Tuple[bytes, str]:
        """
        Export products for this organization.
        Respects filters and optional id list (selected rows).
        Never crosses organization boundaries.
        """
        filters = filters or {}
        product_svc = ProductService(self.db)

        if ids:
            products = [
                product_svc.get_product(pid, organization_id)
                for pid in ids
                if self.repo.get_by_id_safe(pid, organization_id)
            ]
        else:
            result = product_svc.list_products(
                organization_id=organization_id,
                per_page=50000,
                page=1,
                **{k: v for k, v in filters.items() if k in {
                    "search_term", "category_id", "product_type", "status", "currency",
                }},
            )
            products = result.get("items", [])

        headers = [
            "Name", "SKU / Code", "Type", "Category", "Description",
            "Unit", "Currency", "Unit Price", "Default Discount %",
            "Country", "GST/VAT Group", "Invoice Description", "Status",
            "Billing Frequency", "Brand", "Created At",
        ]

        def _get_category_name(cat_id: Optional[int]) -> str:
            if not cat_id:
                return ""
            try:
                cat = self.cat_repo.get_by_id(cat_id, organization_id)
                return cat.name if cat else ""
            except Exception:
                return ""

        rows = []
        for p in products:
            status_val = "active" if getattr(p, "is_active", True) else "inactive"
            if getattr(p, "deleted_at", None):
                status_val = "archived"
            rows.append([
                p.name or "",
                p.code or "",
                p.product_type or "",
                _get_category_name(p.category_id),
                p.description or "",
                getattr(p, "unit_label", "") or "",
                p.currency or "",
                str(p.default_price or 0),
                str(getattr(p, "default_discount", 0) or 0),
                getattr(p, "country", "") or "",
                getattr(p, "gst_vat_group", "") or "",
                getattr(p, "invoice_description", "") or "",
                status_val,
                getattr(p, "billing_frequency", "") or "",
                getattr(p, "brand", "") or "",
                str(p.created_at.date() if p.created_at else ""),
            ])

        if fmt == "xlsx":
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Products"

            hfill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF")
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = hfont
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(ci)].width = max(15, len(h) + 4)

            for r in rows:
                ws.append(r)

            ws.freeze_panes = "A2"
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else:  # csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(rows)
            return output.getvalue().encode("utf-8"), "text/csv"

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _get_org_category_map(self, organization_id: int) -> Dict[str, ProductCategory]:
        """Return {name.lower(): category} for all non-deleted org categories."""
        try:
            cats = self.cat_repo.list_all(organization_id)
            return {c.name.lower(): c for c in cats if c and c.name}
        except Exception:
            return {}

    def _generate_category_code(self, name: str, organization_id: int) -> str:
        """Generate a unique category code from a name (mirrors _make_unique_code_and_name)."""
        base = re.sub(r"[^A-Za-z0-9]+", "-", name.strip()).strip("-").upper()[:40] or "CAT"
        code = base
        suffix = 1
        while self.cat_repo.exists(organization_id, code=code):
            suffix += 1
            code = f"{base}-{suffix}"
        return code

    def _resolve_or_create_category(
        self,
        name: str,
        organization_id: int,
        existing_categories: Dict[str, ProductCategory],
        auto_create: bool,
    ) -> Tuple[Optional[int], List[str]]:
        """
        Resolve a category name to an ID.
        - If it exists, return its ID.
        - If not and auto_create=True, create it and return new ID.
        - If not and auto_create=False, return None + validation error.
        """
        if not name or not name.strip():
            return None, []
        key = name.strip().lower()
        if key in existing_categories:
            return existing_categories[key].id, []
        if auto_create:
            try:
                new_cat = self.cat_repo.create(
                    organization_id,
                    name=name.strip(),
                    code=self._generate_category_code(name, organization_id),
                    is_active=True,
                )
                existing_categories[key] = new_cat  # update in-place for subsequent rows
                logger.info(f"[import] Auto-created category '{name}' for org {organization_id}")
                return new_cat.id, [f"Category '{name}' was auto-created."]
            except Exception as exc:
                return None, [f"Failed to auto-create category '{name}': {exc}"]
        else:
            return None, [f"Category '{name}' does not exist in this organization."]

    def _map_and_validate_row(
        self,
        raw: Dict[str, str],
        column_map: Dict[str, str],
        row_index: int,
        organization_id: int,
        org_currency: str,
        existing_categories: Dict[str, ProductCategory],
        auto_create_categories: bool,
    ) -> Tuple[Dict[str, Any], List[str], List[str]]:
        """
        Map raw CSV/XLSX columns to product fields, validate all values.
        Returns (mapped_data, errors, warnings).
        """
        mapped: Dict[str, Any] = {}
        errors: List[str] = []
        warnings: List[str] = []

        for file_col, product_field in column_map.items():
            value = raw.get(file_col, "")
            if value is None:
                value = ""
            value = str(value).strip()
            if product_field in ("category", "tax_category"):
                # handled separately below
                mapped[f"_raw_{product_field}"] = value
            else:
                mapped[product_field] = value

        # --- Required fields ---
        name = mapped.get("name", "")
        if not name:
            errors.append("Row is missing required field: Name")

        code = mapped.get("code", "")
        if not code:
            errors.append("Row is missing required field: SKU / Code")

        # --- Product type ---
        pt = mapped.get("product_type", "service").lower()
        if pt:
            if pt not in VALID_PRODUCT_TYPES:
                errors.append(f"Invalid product type '{pt}'. Accepted: {', '.join(sorted(VALID_PRODUCT_TYPES))}")
            else:
                mapped["product_type"] = pt
        else:
            mapped["product_type"] = "service"
            warnings.append("Product type not specified — defaulted to 'service'")

        # --- Currency ---
        currency = (mapped.get("currency") or "").upper().strip()
        if currency:
            if currency not in VALID_CURRENCIES:
                errors.append(f"Invalid currency '{currency}'. Use 3-letter ISO codes like USD, EUR, INR.")
            else:
                mapped["currency"] = currency
        else:
            mapped["currency"] = org_currency

        # --- Numeric fields ---
        for field_name, display_name in [
            ("default_price", "Unit Price"),
            ("default_discount", "Default Discount"),
            ("original_price", "Original Price"),
            ("tax_percentage", "Tax Rate"),
        ]:
            raw_val = mapped.get(field_name, "")
            if raw_val:
                numeric = _normalize_decimal(str(raw_val))
                if numeric is None:
                    errors.append(f"Invalid numeric value for '{display_name}': '{raw_val}'")
                else:
                    if field_name == "default_discount" and not (0 <= numeric <= 100):
                        errors.append(f"Default Discount must be between 0 and 100. Got: {numeric}")
                    elif field_name in ("default_price", "original_price", "tax_percentage") and numeric < 0:
                        errors.append(f"'{display_name}' cannot be negative. Got: {numeric}")
                    else:
                        mapped[field_name] = numeric
            else:
                if field_name == "default_price":
                    mapped[field_name] = 0.0
                else:
                    mapped.pop(field_name, None)

        # --- Status ---
        status_raw = (mapped.get("status") or "").strip()
        if status_raw:
            is_active = _normalize_status(status_raw)
            if is_active is None:
                warnings.append(f"Unrecognized status '{status_raw}' — defaulted to active")
                mapped["is_active"] = True
            else:
                mapped["is_active"] = is_active
        else:
            mapped["is_active"] = True
        mapped.pop("status", None)

        # --- Billing frequency ---
        bf = (mapped.get("billing_frequency") or "").lower().strip()
        if bf:
            if bf not in VALID_BILLING_FREQUENCIES:
                warnings.append(f"Unrecognized billing frequency '{bf}' — defaulted to 'one_time'")
                mapped["billing_frequency"] = "one_time"
            else:
                mapped["billing_frequency"] = bf
        else:
            # auto-derive from type
            type_freq_map = {
                "good": "one_time", "service": "one_time",
                "subscription": "monthly", "usage": "usage_based",
                "retainer": "monthly",
            }
            mapped["billing_frequency"] = type_freq_map.get(mapped.get("product_type", "service"), "one_time")

        # --- Category resolution ---
        cat_name = mapped.pop("_raw_category", "")
        if cat_name:
            cat_id, cat_warns = self._resolve_or_create_category(
                cat_name, organization_id, existing_categories, auto_create_categories,
            )
            if cat_id is None and not auto_create_categories:
                errors.append(f"Category '{cat_name}' not found and auto-create is disabled.")
            elif cat_id:
                mapped["category_id"] = cat_id
                warnings.extend(cat_warns)
            else:
                warnings.append(f"Category '{cat_name}' could not be created; product imported without category")
        
        # --- Tax category (validate only — don't auto-create tax categories) ---
        tax_cat_name = mapped.pop("_raw_tax_category", "")
        if tax_cat_name:
            warnings.append(
                f"Tax category '{tax_cat_name}' in import file — "
                "please assign via the Tax module after import."
            )

        # Remove any non-product fields that leaked through
        for extra in ["_raw_category", "_raw_tax_category", "status"]:
            mapped.pop(extra, None)

        # Only keep allowed fields
        cleaned = {k: v for k, v in mapped.items() if k in PRODUCT_ALLOWED_FIELDS or k == "category_id"}
        # category_id is not in PRODUCT_ALLOWED_FIELDS (it's handled separately by service) — keep it
        cleaned["name"] = name
        cleaned["code"] = code

        return cleaned, errors, warnings

    def _check_duplicate(
        self,
        organization_id: int,
        code: Optional[str],
        name: Optional[str],
    ) -> Tuple[bool, Optional[int], Optional[str]]:
        """Check if a product with the same code or name already exists."""
        if code:
            existing = self.repo.get_by_code(organization_id, code)
            if existing:
                return True, existing.id, existing.code
        if name:
            existing = self.repo.get_by_name(organization_id, name)
            if existing:
                return True, existing.id, existing.code
        return False, None, None

    def _create_product_from_mapped(
        self,
        svc: ProductService,
        organization_id: int,
        user_id: int,
        mapped: Dict[str, Any],
        auto_create_categories: bool,
    ) -> Product:
        return svc.create_product(
            organization_id=organization_id,
            created_by=user_id,
            **{k: v for k, v in mapped.items() if v != "" and v is not None},
        )

    def _update_product_from_mapped(
        self,
        svc: ProductService,
        product_id: int,
        organization_id: int,
        user_id: int,
        mapped: Dict[str, Any],
        auto_create_categories: bool,
    ) -> Product:
        return svc.update_product(
            product_id=product_id,
            organization_id=organization_id,
            updated_by=user_id,
            **{k: v for k, v in mapped.items() if k not in {"name", "code"}},
        )

    def _make_unique_code_and_name(
        self, mapped: Dict[str, Any], organization_id: int
    ) -> Dict[str, Any]:
        """Generate a unique code+name for 'create_copy' duplicate strategy."""
        base_code = f"{mapped.get('code', 'COPY')}-COPY"
        code = base_code
        suffix = 1
        while self.repo.exists(organization_id, code=code):
            suffix += 1
            code = f"{base_code}-{suffix}"

        base_name = f"{mapped.get('name', 'Copy')} Copy"
        name = base_name
        suffix = 1
        while self.repo.get_by_name(organization_id, name):
            suffix += 1
            name = f"{base_name} {suffix}"

        return {**mapped, "code": code, "name": name}
