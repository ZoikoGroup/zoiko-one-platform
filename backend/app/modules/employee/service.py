import csv
import hashlib
import io
import logging
import os
import re
import secrets
import string
import tempfile
from datetime import date, datetime, timedelta
from typing import Optional, List
from decimal import Decimal

from sqlalchemy import cast, extract, func, Integer, text
from sqlalchemy.orm import Session

from app.database import Base

logger = logging.getLogger("zoiko.employee.service")

from app.modules.employee.models import (
    Employee, EmploymentType, EmployeeStatus, UserRole, Gender,
    SecurityActionPurpose, SecurityActionToken,
)
from app.modules.employee.schema import (
    EmployeeCreate, EmployeeUpdate,
    UserCreateRequest, UserUpdateRequest,
    LoginRequest, RegisterRequest,
    ChangeManagerRequest, ConfirmProbationRequest,
    PromoteEmployeeRequest, TransferEmployeeRequest,
    ResignationRequest, ExitEmployeeRequest, EmployeeExportRequest,
    EmployeeCompensationCreate, EmployeeCompensationUpdate,
    EmployeeBenefitCreate,
)
from app.modules.hr.models import (
    Organization, OrganizationStatus, Department, Designation,
    EmployeeProfile, EmployeeReporting, EmployeeLifecycle, EmployeeHistory,
    EmployeeCompensation, EmployeeBenefit, Allowance, CompensationItem,
    Asset, LeaveRequest, LeaveBalance, AttendanceRecord, ShiftRoster,
    EssRequest, TravelRequest, TravelExpense, ComplianceRecord,
    PerformanceReview, PerformanceGoal, PerformanceKpi, PerformanceFeedback, Appraisal,
    LearningEnrollment, LearningCertification, LearningSkill, LearningQuizAttempt,
    LearningTrainingProgramAssignment, WfSuccession, HrDocument, DocumentAssignment,
    OnboardingNewHire,
    SalaryRevision, TravelApproval, TravelReceipt,
    RecruitmentInterviewFeedback, RecruitmentOfferApproval,
    Holiday, AssetMaintenanceRequest, AssetRequest, AssetReport, EngagementSurvey,
    OnboardingPreboardingTask, RecruitmentInterview, RecruitmentDocument,
    WfPlan, WfHeadcount, WfReport,
    LearningCourse, LearningPath, LearningAssessment, LearningTrainingProgram,
    LearningCalendarEvent, HrDocumentVersion, DocumentApprovalStep, DocumentApprovalLog,
)
from app.modules.time.models import TimeEntry
from app.modules.comply.models import PolicyAcknowledgement, CompliancePolicy
from app.modules.insights.models import Report, ReportRun
from app.modules.payroll.models import PayrollRun, ComplianceDocument, PayrollActivityLog
from app.modules.super_admin.models import (
    PlatformProduct, OrganizationProduct, ProductStatus,
    AuditLog, Notification, SupportTicket, SecurityEvent, ApprovalHistory, LoginActivity,
)
from app.modules.billing.models import (
    BillingSetting, BillingCustomer, CustomerContact, ProductCategory, Product,
    PricingPlan, PriceList, PriceListItem, PricingRule, Discount, CurrencyPricing,
    TaxPricing, TaxGroup, Contract, ContractAmendment, Quotation, SubscriptionPlan,
    Subscription, SubscriptionEvent, Invoice, InvoiceStatusHistory, PaymentMethod,
    Payment, PaymentAllocation, CreditNote, CreditNoteApplication, Refund, TaxRate, Tax,
    DunningCase, CollectionsCase, CollectionAction, RevenueRecognitionSchedule,
    BillingAuditLog, BillingConfiguration, CustomerDocument, CustomerNote,
)
from app.modules.payroll.models import PayrollEmployee
from app.core.security import hash_password, verify_password, create_access_token
from app.core.exceptions import (
    NotFoundException, AlreadyExistsException,
    UnauthorizedException, BadRequestException,
)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER
# ═══════════════════════════════════════════════════════════════════════════════

def derive_employee_id_prefix(org_name: str) -> str:
    """Derive a 2-letter employee-ID prefix from an organization name.

    Rules:
    - Strip all non-alpha characters from org_name.
    - Take the first two letters, uppercased.
    - If fewer than 2 alpha chars exist, pad with 'X' (e.g. "A1" -> "AX").
    - If no alpha chars at all, fall back to "OR".
    """
    alpha_only = re.sub(r"[^A-Za-z]", "", org_name or "")
    if len(alpha_only) >= 2:
        return alpha_only[:2].upper()
    if len(alpha_only) == 1:
        return (alpha_only + "X").upper()
    return "OR"


def _generate_employee_id(db: Session, organization_id: int) -> str:
    """Generate an org-scoped employee_id like ZO0001, AC0002, etc.

    Concurrency-safe via pg_advisory_xact_lock on (organization_id).
    Filters by LIKE '<prefix>%' so historical EMP####-style IDs are untouched.
    """
    from app.modules.hr.models import Organization

    db.execute(
        text("SELECT pg_advisory_xact_lock(:org_key)"),
        {"org_key": organization_id + 7000000},
    )

    org = db.query(Organization).filter(Organization.id == organization_id).first()
    prefix = org.employee_id_prefix if org and org.employee_id_prefix else "OR"

    like_pattern = f"{prefix}%"
    max_seq = (
        db.query(Employee.employee_id)
        .filter(
            Employee.organization_id == organization_id,
            Employee.employee_id.isnot(None),
            Employee.employee_id.like(like_pattern),
        )
        .all()
    )

    existing_nums = []
    for (eid,) in max_seq:
        num_part = eid[len(prefix):]
        if num_part.isdigit():
            existing_nums.append(int(num_part))

    next_num = max(existing_nums) + 1 if existing_nums else 1
    return f"{prefix}{next_num:04d}"



def _generate_temp_password(length: int = 12) -> str:
    chars = string.ascii_letters + string.digits
    return ''.join(secrets.choice(chars) for _ in range(length))


def _role_to_default_title(role: UserRole) -> str:
    titles = {
        UserRole.SUPER_ADMIN: "Super Administrator",
        UserRole.ADMIN: "Organization Administrator",
        UserRole.HR_ADMIN: "HR Administrator",
        UserRole.BILLING_ADMIN: "Billing Administrator",
        UserRole.EMPLOYEE: "Employee",
    }
    return titles.get(role, "Employee")


def _full_name(employee) -> str:
    return f"{employee.first_name} {employee.last_name}".strip() or employee.email


def _notify_email(sender_name: str, **kwargs) -> bool:
    """Best-effort outbound email — never blocks the underlying action."""
    try:
        from app.services import email_service
        sender_fn = getattr(email_service, sender_name)
        sender_fn(**kwargs)
        return True
    except Exception:
        logger.exception("Failed to send email via %s", sender_name)
        return False


# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY ACTION TOKENS (single-use, expiring)
# ═══════════════════════════════════════════════════════════════════════════════

TOKEN_TTL_HOURS = 24
TOKEN_TIMEZONE = "UTC"
INVALID_TOKEN_MESSAGE = "This link is invalid or has expired. Please request a new one."


def _token_hash(raw_token: str) -> str:
    return hashlib.sha256(raw_token.encode("utf-8")).hexdigest()


def _format_token_datetime(dt: datetime) -> str:
    return dt.strftime("%b %d, %Y at %I:%M %p")


def _org_workspace_name(db: Session, organization_id) -> str:
    if not organization_id:
        return ""
    from app.modules.hr.models import Organization
    org = db.query(Organization).filter(Organization.id == organization_id).first()
    return (org.organization_name or org.display_name or "") if org else ""


def _issue_action_token(db: Session, email: str, organization_id, purpose) -> tuple[str, datetime]:
    """Create a single-use token row; returns (raw_token, expires_at). Only the
    SHA-256 hash is stored — the raw token is embedded in the emailed link."""
    raw_token = secrets.token_urlsafe(32)
    expires_at = datetime.utcnow() + timedelta(hours=TOKEN_TTL_HOURS)
    token = SecurityActionToken(
        email=email,
        organization_id=organization_id,
        purpose=purpose,
        token_hash=_token_hash(raw_token),
        expires_at=expires_at,
    )
    db.add(token)
    db.flush()
    return raw_token, expires_at


def _action_link(purpose, raw_token: str) -> str:
    base = os.environ.get("API_BASE_URL", "http://localhost:8000")
    path = "accept-invite" if purpose == SecurityActionPurpose.INVITE else "reset-password"
    return f"{base}/auth/{path}?token={raw_token}"


def _consume_action_token(db: Session, raw_token: str, purpose) -> Optional[dict]:
    """Atomically consume a single-use action token.

    Single UPDATE...RETURNING statement — no SELECT-then-UPDATE window, so two
    concurrent requests can never both succeed. The purpose filter blocks
    cross-use (invite tokens cannot reset passwords and vice versa). Returns
    {"email": ..., "organization_id": ...} or None when the token is unknown,
    expired, already used, or for the wrong purpose.
    """
    row = db.execute(
        text(
            """
            UPDATE security_action_tokens
            SET used_at = CURRENT_TIMESTAMP
            WHERE token_hash = :hash
              AND purpose = :purpose
              AND used_at IS NULL
              AND expires_at > :now
            RETURNING email, organization_id
            """
        ),
        {"hash": _token_hash(raw_token), "purpose": purpose.name, "now": datetime.utcnow()},
    ).fetchone()
    if row is None:
        return None
    return {"email": row[0], "organization_id": row[1]}


def validate_action_token(db: Session, raw_token: str, purpose) -> Optional[dict]:
    """Read-only validity check for the GET page. Returns a render context or
    None for EVERY invalid state (unknown, used, expired, wrong purpose) so the
    GET page cannot distinguish them.
    """
    row = db.execute(
        text(
            """
            SELECT email, organization_id, purpose, expires_at, used_at
            FROM security_action_tokens
            WHERE token_hash = :hash
            """
        ),
        {"hash": _token_hash(raw_token)},
    ).fetchone()
    if row is None:
        return None
    email, organization_id, purpose_stored, expires_at, used_at = row
    if isinstance(expires_at, str):
        try:
            expires_at = datetime.fromisoformat(expires_at)
        except ValueError:
            return None
    if (
        used_at is not None
        or purpose_stored != purpose.name
        or expires_at <= datetime.utcnow()
    ):
        return None
    employee = db.query(Employee).filter(Employee.email == email).first()
    return {
        "token": raw_token,
        "email": email,
        "organization_id": organization_id,
        "first_name": (employee.first_name or "") if employee else "",
        "workspace_name": _org_workspace_name(db, organization_id),
    }


def complete_action_token(db: Session, raw_token: str, purpose, new_password: str) -> dict:
    """Consume the token and set the new password in ONE transaction. Consuming
    first and only then mutating the employee means any failure rolls the used_at
    flag back with it. Raises BadRequestException (generic) for every invalid
    state — the same message whether the token never existed, expired, or was used.
    """
    consumed = _consume_action_token(db, raw_token, purpose)
    if consumed is None:
        raise BadRequestException(INVALID_TOKEN_MESSAGE)

    employee = db.query(Employee).filter(Employee.email == consumed["email"]).first()
    if not employee:
        raise BadRequestException(INVALID_TOKEN_MESSAGE)

    employee.hashed_password = hash_password(new_password)
    db.commit()
    db.refresh(employee)

    _notify_email(
        "send_org_admin_password_changed_email",
        email=employee.email,
        first_name=employee.first_name or _full_name(employee),
        event_time_local=_format_token_datetime(datetime.utcnow()),
        timezone=TOKEN_TIMEZONE,
        organization_id=employee.organization_id,
        db=db,
    )
    return {"message": "Password set successfully. You can now sign in."}


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH SERVICE
# ═══════════════════════════════════════════════════════════════════════════════

def login_employee(db: Session, data: LoginRequest) -> dict:
    employee = db.query(Employee).filter(Employee.email == data.email).first()
    if not employee:
        raise UnauthorizedException("Invalid email or password.")

    if not verify_password(data.password, employee.hashed_password):
        raise UnauthorizedException("Invalid email or password.")

    if employee.organization_id:
        org = db.query(Organization).filter(Organization.id == employee.organization_id).first()
        if org:
            if org.status == OrganizationStatus.PENDING:
                raise UnauthorizedException(
                    "Your organization registration is awaiting Super Admin approval. "
                    "You will be able to sign in after approval."
                )
            elif org.status == OrganizationStatus.REJECTED:
                reason = f" Reason: {org.rejection_reason}" if org.rejection_reason else ""
                raise UnauthorizedException(
                    f"Your organization registration has been rejected.{reason}"
                )
            elif org.status == OrganizationStatus.SUSPENDED:
                raise UnauthorizedException(
                    "Your organization has been suspended. Please contact support."
                )
            elif org.status == OrganizationStatus.DEACTIVATED:
                raise UnauthorizedException(
                    "Your organization has been deactivated. Please contact support."
                )

    if not employee.is_active:
        raise UnauthorizedException("Your account has been deactivated.")

    if employee.status == EmployeeStatus.DEACTIVATED:
        raise UnauthorizedException("Your account has been deactivated.")

    from app.modules.hr.models import Organization as HrOrg
    org_obj = None
    if employee.organization_id:
        org_obj = db.query(HrOrg).filter(HrOrg.id == employee.organization_id).first()
    org_code = org_obj.organization_code if org_obj else None

    token = create_access_token(data={
        "sub":  employee.email,
        "role": employee.role.value,
        "id":   employee.id,
        "organization_id": employee.organization_id,
        "organization_code": org_code,
    })

    refresh_token = create_access_token(
        data={"sub": employee.email, "id": employee.id, "organization_id": employee.organization_id},
        expires_delta=timedelta(days=7),
    )

    # Serialize employee object using Pydantic schema
    from app.modules.employee.schema import EmployeeResponse
    emp_data = EmployeeResponse.model_validate(employee).model_dump()
    # Include the org's enabled products for sidebar filtering
    if employee.organization_id:
        product_rows = db.query(PlatformProduct.code).join(OrganizationProduct).filter(
            OrganizationProduct.organization_id == employee.organization_id,
            OrganizationProduct.is_enabled == True,
        ).all()
        emp_data["products"] = [r[0] for r in product_rows]
        logger.debug("[PRODUCTS] Login: user=%s org_id=%s products=%s", employee.email, employee.organization_id, emp_data['products'])
    else:
        logger.debug("[PRODUCTS] Login: user=%s no organization, products=[]", employee.email)
    employee_serialized = EmployeeResponse.model_validate(emp_data)

    return {
        "access_token": token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "employee": employee_serialized,
    }


_PRODUCT_DISPLAY_NAMES = {
    "hr": "Zoiko HR",
    "time": "ZoikoTime",
    "payroll": "Zoiko Payroll",
    "billing": "Zoiko Billing",
    "projects": "Zoiko Projects",
    "comply": "Zoiko Comply",
    "insights": "Zoiko Insights",
    "spend": "Zoiko Spend",
    "inventory": "Zoiko Inventory",
    "docs": "Zoiko Docs Pro",
}

def _ensure_platform_products(db: Session, codes: list[str]) -> list[PlatformProduct]:
    """Ensure PlatformProduct rows exist for the given codes; create any missing ones."""
    if not codes:
        return []
    existing = db.query(PlatformProduct).filter(
        PlatformProduct.code.in_(codes),
    ).all()
    existing_map = {p.code: p for p in existing}
    missing = [c for c in codes if c not in existing_map]
    if missing:
        logger.debug("[PRODUCTS] Creating missing PlatformProduct records: %s", missing)
    for code in codes:
        if code not in existing_map:
            prod = PlatformProduct(
                code=code,
                name=_PRODUCT_DISPLAY_NAMES.get(code, code.title()),
                description=f"{code} module",
                status=ProductStatus.ACTIVE,
            )
            db.add(prod)
            db.flush()
            existing_map[code] = prod
    result = list(existing_map.values())
    logger.debug("[PRODUCTS] _ensure_platform_products(codes=%s) -> %d products: %s", codes, len(result), [p.code for p in result])
    return result

def _save_org_products(db: Session, org_id: int, product_codes) -> None:
    logger.debug("[PRODUCTS] _save_org_products(org_id=%s, product_codes=%s)", org_id, product_codes)
    if isinstance(product_codes, str):
        product_codes = [product_codes]
    if not product_codes:
        logger.debug("[PRODUCTS] No product_codes provided — skipping OrganizationProduct creation (approve flow will handle orgs with 0 products)")
        return
    if "all" in product_codes:
        products = db.query(PlatformProduct).filter(
            PlatformProduct.status == ProductStatus.ACTIVE,
        ).all()
        logger.debug("[PRODUCTS] 'all' specified, assigning ALL active: %s", [p.code for p in products])
    else:
        products = _ensure_platform_products(db, product_codes)

    from app.core.code_generation import generate_tenant_code

    created = 0
    for prod in products:
        existing = db.query(OrganizationProduct).filter(
            OrganizationProduct.organization_id == org_id,
            OrganizationProduct.product_id == prod.id,
        ).first()
        if not existing:
            tenant_code = generate_tenant_code(db, org_id, prod.code)
            db.add(OrganizationProduct(
                organization_id=org_id,
                product_id=prod.id,
                tenant_code=tenant_code,
                is_enabled=True,
            ))
            created += 1
    logger.debug("[PRODUCTS] _save_org_products: %d new OrganizationProduct records created for org_id=%s", created, org_id)
    # Verify what was saved
    all_saved = db.query(OrganizationProduct).filter(
        OrganizationProduct.organization_id == org_id,
    ).all()
    saved_codes = []
    for op in all_saved:
        pp = db.query(PlatformProduct).filter(PlatformProduct.id == op.product_id).first()
        if pp:
            saved_codes.append(pp.code)
    logger.debug("[PRODUCTS] Verification: org_id=%s now has %d products: %s", org_id, len(saved_codes), saved_codes)

# TODO: This function is duplicated in hr/service.py.  Changes here must be
# mirrored there, or the two copies should be consolidated into one.
def register_enterprise(db: Session, data: RegisterRequest) -> dict:
    existing = db.query(Employee).filter(Employee.email == data.email).first()
    if existing:
        raise AlreadyExistsException("Employee", "email")

    from app.core.code_generation import generate_organization_code, generate_uuid, generate_employee_code

    org_code = generate_organization_code(data.organization, db)
    org_uuid = generate_uuid()

    # Also ensure legacy `code` is set (backward compat)
    legacy_code = data.organization[:50].upper().replace(" ", "_")
    suffix = 1
    while db.query(Organization).filter(Organization.organization_code == legacy_code).first():
        legacy_code = f"{data.organization[:45].upper().replace(' ', '_')}_{suffix}"
        suffix += 1

    org = Organization(
        name=data.organization,
        code=legacy_code,
        uuid=org_uuid,
        organization_code=org_code,
        organization_name=data.organization,
        status=OrganizationStatus.PENDING,
        address=data.address,
        city=data.city,
        state=data.state,
        country=data.country,
        timezone=data.timezone or "UTC",
        industry=data.industry,
        employee_id_prefix=derive_employee_id_prefix(data.organization),
    )
    db.add(org)
    db.commit()
    db.refresh(org)

    dept_code = f"MGMT_{org.id}"
    dept_department_code = f"{org_code}DEP001"
    dept = Department(name="Management", code=dept_code, department_code=dept_department_code, description="Company management", organization_id=org.id)
    db.add(dept)
    db.commit()
    db.refresh(dept)

    name_parts = data.name.strip().split(" ", 1)
    first_name = name_parts[0]
    last_name = name_parts[1] if len(name_parts) > 1 else "Admin"

    employee_code = generate_employee_code(db, organization_id=org.id)

    employee = Employee(
        email=data.email,
        hashed_password=hash_password(data.password),
        role=UserRole.ADMIN,
        is_active=True,
        first_name=first_name,
        last_name=last_name,
        phone="",
        employee_code=employee_code,
        job_title="System Administrator",
        employment_type=EmploymentType.FULL_TIME,
        status=EmployeeStatus.ACTIVE,
        date_of_joining=date.today(),
        department_id=dept.id,
        organization_id=org.id,
    )
    db.add(employee)
    db.commit()
    db.refresh(employee)

    from app.modules.super_admin.models import AuditLog, AuditAction, Notification
    audit = AuditLog(
        action=AuditAction.CREATE,
        entity_type="Organization",
        entity_id=org.id,
        performed_by=employee.id,
        performed_by_email=employee.email,
        details={"organization": org.name, "code": org.code, "status": "PENDING"},
    )
    db.add(audit)

    notification = Notification(
        title="New Organization Registration",
        message=f"Organization '{org.name}' has registered and is awaiting approval.",
        notification_type="org_registration",
        priority="high",
        target_org_id=org.id,
        target_user_id=employee.id,
    )
    db.add(notification)

    # Validate and save product selection
    selected_products = data.products or ([data.product] if data.product else None)
    if not selected_products:
        raise BadRequestException("At least one product must be selected during registration.")
    logger.debug("[PRODUCTS] Register: org='%s' data.products=%s data.product=%s -> selected_products=%s", data.organization, data.products, data.product, selected_products)
    _save_org_products(db, org.id, selected_products)

    db.commit()

    # Send registration received email (non-blocking)
    from app.services.email_service import send_registration_received
    try:
        send_registration_received(data.email, org.name, db=db)
    except Exception as e:
        logger.warning(f"[email] Failed to send registration email to admin {data.email}: {e}")

    if data.registered_email and data.registered_email != data.email:
        try:
            send_registration_received(data.registered_email, org.name, db=db)
        except Exception as e:
            logger.warning(f"[email] Failed to send registration email to {data.registered_email}: {e}")

    return {
        "message": "Organization registered successfully. Awaiting Super Admin approval.",
        "organization_id": org.id,
        "organization_name": org.name,
    }


def change_password(
    db: Session,
    employee_id: int,
    current_password: str,
    new_password: str,
) -> dict:
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise NotFoundException("Employee", employee_id)

    if not verify_password(current_password, employee.hashed_password):
        raise UnauthorizedException("Current password is incorrect.")

    employee.hashed_password = hash_password(new_password)
    db.commit()
    db.refresh(employee)

    if employee.role == UserRole.ADMIN:
        _notify_email(
            "send_org_admin_password_changed_email",
            email=employee.email,
            first_name=employee.first_name or _full_name(employee),
            event_time_local=_format_token_datetime(datetime.utcnow()),
            timezone=TOKEN_TIMEZONE,
            organization_id=employee.organization_id,
            db=db,
        )

    return {"message": "Password changed successfully."}


# ═══════════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT SERVICE (Organization Admin)
# ═══════════════════════════════════════════════════════════════════════════════

def create_organization_user(
    db: Session,
    data: "UserCreateRequest",
    organization_id: int,
    created_by_id: int,
) -> Employee:
    existing = db.query(Employee).filter(Employee.email == data.email).first()
    if existing:
        raise AlreadyExistsException("User", "email")

    temp_password = _generate_temp_password()
    role = data.role

    from app.core.code_generation import generate_employee_code
    new_employee_code = generate_employee_code(db, organization_id=organization_id)

    employee = Employee(
        email=data.email,
        hashed_password=hash_password(temp_password),
        employee_code=new_employee_code,
        employee_id=_generate_employee_id(db, organization_id=organization_id),
        role=role,
        is_active=True,
        first_name=data.first_name,
        last_name=data.last_name,
        phone=data.phone or "",
        job_title=data.job_title or _role_to_default_title(role),
        employment_type=EmploymentType.FULL_TIME,
        status=EmployeeStatus.ACTIVE,
        date_of_joining=date.today(),
        organization_id=organization_id,
        created_by=created_by_id,
    )
    db.add(employee)
    db.commit()
    db.refresh(employee)

    if role == UserRole.ADMIN:
        raw_token, expires_at = _issue_action_token(db, employee.email, employee.organization_id, SecurityActionPurpose.INVITE)
        db.commit()
        inviter = db.query(Employee).filter(Employee.id == created_by_id).first() if created_by_id else None
        inviter_name = _full_name(inviter) if inviter else ""
        _notify_email(
            "send_org_admin_invite_email",
            email=employee.email,
            first_name=employee.first_name or _full_name(employee),
            inviter_name=inviter_name,
            workspace_name=_org_workspace_name(db, employee.organization_id),
            expires_at_local=_format_token_datetime(expires_at),
            timezone=TOKEN_TIMEZONE,
            action_url=_action_link(SecurityActionPurpose.INVITE, raw_token),
            organization_id=employee.organization_id,
            db=db,
        )
    else:
        _notify_email(
            "send_employee_welcome_email",
            email=employee.email,
            employee_name=_full_name(employee),
            temporary_password=temp_password,
            organization_id=employee.organization_id,
            db=db,
        )

    return employee, temp_password


def get_organization_users(
    db: Session,
    organization_id: int,
    search: Optional[str] = None,
    role: Optional[UserRole] = None,
    status: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(Employee).filter(Employee.organization_id == organization_id)

    if search:
        term = f"%{search}%"
        query = query.filter(
            (Employee.first_name.ilike(term)) |
            (Employee.last_name.ilike(term)) |
            (Employee.email.ilike(term)) |
            (Employee.employee_id.ilike(term)) |
            (Employee.employee_code.ilike(term))
        )

    if role:
        query = query.filter(Employee.role == role)

    if status:
        if status == "active":
            query = query.filter(Employee.is_active == True)
        elif status == "inactive":
            query = query.filter(Employee.is_active == False)

    total = query.count()
    users = query.order_by(Employee.created_at.desc()).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    return {"total": total, "page": page, "per_page": per_page, "items": users}


def get_organization_user(
    db: Session,
    user_id: int,
    organization_id: int,
) -> Employee:
    user = db.query(Employee).filter(
        Employee.id == user_id,
        Employee.organization_id == organization_id,
    ).first()
    if not user:
        raise NotFoundException("User", user_id)
    return user


def update_organization_user(
    db: Session,
    user_id: int,
    data: "UserUpdateRequest",
    organization_id: int,
    updated_by_id: int,
) -> Employee:
    user = get_organization_user(db, user_id, organization_id)
    old_role = user.role
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(user, field, value)
    user.updated_by = updated_by_id
    db.commit()
    db.refresh(user)

    if "role" in update_data and user.role != old_role:
        if old_role == UserRole.ADMIN or user.role == UserRole.ADMIN:
            _notify_email(
                "send_org_admin_access_changed_email",
                email=user.email,
                first_name=user.first_name or _full_name(user),
                workspace_name=_org_workspace_name(db, user.organization_id),
                effective_date_local=date.today().strftime("%b %d, %Y"),
                organization_id=user.organization_id,
                db=db,
            )

    return user


def deactivate_organization_user(
    db: Session,
    user_id: int,
    organization_id: int,
    updated_by_id: int,
) -> Employee:
    user = get_organization_user(db, user_id, organization_id)
    user.is_active = False
    user.status = EmployeeStatus.INACTIVE
    user.updated_by = updated_by_id
    db.commit()
    db.refresh(user)

    if user.role == UserRole.ADMIN:
        _notify_email(
            "send_org_admin_access_removed_email",
            email=user.email,
            first_name=user.first_name or _full_name(user),
            workspace_name=_org_workspace_name(db, user.organization_id),
            effective_date_local=date.today().strftime("%b %d, %Y"),
            organization_id=organization_id,
            db=db,
        )
    else:
        _notify_email(
            "send_employee_account_status_email",
            email=user.email,
            employee_name=_full_name(user),
            status="deactivated",
            organization_id=organization_id,
            db=db,
        )

    return user


def activate_organization_user(
    db: Session,
    user_id: int,
    organization_id: int,
    updated_by_id: int,
) -> Employee:
    user = get_organization_user(db, user_id, organization_id)
    user.is_active = True
    user.status = EmployeeStatus.ACTIVE
    user.updated_by = updated_by_id
    db.commit()
    db.refresh(user)

    if user.role == UserRole.ADMIN:
        _notify_email(
            "send_org_admin_account_activated_email",
            email=user.email,
            first_name=user.first_name or _full_name(user),
            workspace_name=_org_workspace_name(db, user.organization_id),
            organization_id=organization_id,
            db=db,
        )
    else:
        _notify_email(
            "send_employee_account_status_email",
            email=user.email,
            employee_name=_full_name(user),
            status="activated",
            organization_id=organization_id,
            db=db,
        )

    return user


def suspend_organization_user(
    db: Session,
    user_id: int,
    organization_id: int,
    updated_by_id: int,
) -> Employee:
    user = get_organization_user(db, user_id, organization_id)
    user.is_active = False
    user.status = EmployeeStatus.SUSPENDED
    user.updated_by = updated_by_id
    db.commit()
    db.refresh(user)

    if user.role == UserRole.ADMIN:
        _notify_email(
            "send_org_admin_account_locked_email",
            email=user.email,
            first_name=user.first_name or _full_name(user),
            organization_id=organization_id,
            db=db,
        )
    else:
        _notify_email(
            "send_employee_account_status_email",
            email=user.email,
            employee_name=_full_name(user),
            status="suspended",
            organization_id=organization_id,
            db=db,
        )

    return user


def archive_organization_user(
    db: Session,
    user_id: int,
    organization_id: int,
    updated_by_id: int,
) -> Employee:
    user = get_organization_user(db, user_id, organization_id)
    user.is_active = False
    user.status = EmployeeStatus.ARCHIVED
    user.updated_by = updated_by_id
    db.commit()
    db.refresh(user)

    if user.role == UserRole.ADMIN:
        _notify_email(
            "send_org_admin_access_removed_email",
            email=user.email,
            first_name=user.first_name or _full_name(user),
            workspace_name=_org_workspace_name(db, user.organization_id),
            effective_date_local=date.today().strftime("%b %d, %Y"),
            organization_id=organization_id,
            db=db,
        )
    else:
        _notify_email(
            "send_employee_account_status_email",
            email=user.email,
            employee_name=_full_name(user),
            status="archived",
            organization_id=organization_id,
            db=db,
        )

    return user


def request_password_reset(db: Session, email: str) -> dict:
    """Public forgot-password flow for org admins. Issues a single-use RESET
    token and emails a link. Always returns the same generic message whether or
    not the email belongs to an active org admin — never discloses account
    existence (no user enumeration)."""
    generic_message = (
        "If an account exists for that email, a password reset link has been sent."
    )
    employee = db.query(Employee).filter(Employee.email == email).first()
    if not employee or employee.role != UserRole.ADMIN or not employee.is_active:
        return {"message": generic_message}

    raw_token, expires_at = _issue_action_token(
        db, employee.email, employee.organization_id, SecurityActionPurpose.RESET
    )
    db.commit()

    _notify_email(
        "send_org_admin_password_reset_email",
        email=employee.email,
        first_name=employee.first_name or _full_name(employee),
        expires_at_local=_format_token_datetime(expires_at),
        timezone=TOKEN_TIMEZONE,
        action_url=_action_link(SecurityActionPurpose.RESET, raw_token),
        organization_id=employee.organization_id,
        db=db,
    )
    return {"message": generic_message}


def reset_user_password(
    db: Session,
    user_id: int,
    organization_id: int,
    updated_by_id: int,
) -> tuple[Employee, Optional[str]]:
    user = get_organization_user(db, user_id, organization_id)

    if user.role == UserRole.ADMIN:
        raw_token, expires_at = _issue_action_token(db, user.email, user.organization_id, SecurityActionPurpose.RESET)
        user.updated_by = updated_by_id
        db.commit()
        _notify_email(
            "send_org_admin_password_reset_email",
            email=user.email,
            first_name=user.first_name or _full_name(user),
            expires_at_local=_format_token_datetime(expires_at),
            timezone=TOKEN_TIMEZONE,
            action_url=_action_link(SecurityActionPurpose.RESET, raw_token),
            organization_id=organization_id,
            db=db,
        )
        return user, None

    temp_password = _generate_temp_password()
    user.hashed_password = hash_password(temp_password)
    user.updated_by = updated_by_id
    db.commit()
    db.refresh(user)

    _notify_email(
        "send_password_reset",
        email=user.email,
        temp_password=temp_password,
        first_name=user.first_name or _full_name(user),
        organization_id=organization_id,
        db=db,
    )

    return user, temp_password


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE CRUD
# ═══════════════════════════════════════════════════════════════════════════════

def create_employee(db: Session, data: EmployeeCreate, organization_id: Optional[int] = None) -> Employee:
    existing = db.query(Employee).filter(Employee.email == data.email).first()
    if existing:
        raise AlreadyExistsException("Employee", "email")

    if data.department_id:
        dept = db.query(Department).filter(Department.id == data.department_id).first()
        if not dept:
            raise NotFoundException("Department", data.department_id)

    from app.core.code_generation import generate_employee_code

    employee_data = data.model_dump(exclude={"password"})
    employee_data.pop("employee_id", None)
    resolved_org_id = organization_id or employee_data.get("organization_id")
    if not resolved_org_id:
        raise BadRequestException("organization_id is required to create an employee")
    employee = Employee(
        **employee_data,
        hashed_password=hash_password(data.password),
        employee_code=generate_employee_code(db, organization_id=resolved_org_id),
        organization_id=resolved_org_id,
    )

    db.add(employee)
    db.commit()
    db.refresh(employee)

    _notify_email(
        "send_employee_welcome_email",
        email=employee.email,
        employee_name=_full_name(employee),
        temporary_password=data.password,
        organization_id=resolved_org_id,
        db=db,
    )

    return employee


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE IMPORT
# ═══════════════════════════════════════════════════════════════════════════════

_COLUMN_MAP = {
    "employee id": "employee_id",
    "employee_id": "employee_id",
    "employee code": "employee_code",
    "employee_code": "employee_code",
    "first name": "first_name",
    "first_name": "first_name",
    "last name": "last_name",
    "last_name": "last_name",
    "email": "email",
    "password": "password",
    "phone": "phone",
    "job title": "job_title",
    "job_title": "job_title",
    "department": "department_name",
    "designation": "designation_name",
    "reporting manager": "reporting_manager",
    "reporting_manager": "reporting_manager",
    "employment type": "employment_type",
    "employment_type": "employment_type",
    "status": "status",
    "date of joining": "date_of_joining",
    "date_of_joining": "date_of_joining",
    "date of birth": "date_of_birth",
    "date_of_birth": "date_of_birth",
    "gender": "gender",
    "basic salary": "basic_salary",
    "basic_salary": "basic_salary",
    "hra": "hra",
    "hra amount": "hra",
    "hra_amount": "hra",
    "ctc": "ctc",
    "work email": "work_email",
    "work_email": "work_email",
    "personal email": "personal_email",
    "personal_email": "personal_email",
    "confirmation date": "confirmation_date",
    "confirmation_date": "confirmation_date",
    "company": "company",
    "business unit": "business_unit",
    "business_unit": "business_unit",
    "division": "division",
    "team": "team",
    "current address": "current_address",
    "current_address": "current_address",
    "permanent address": "permanent_address",
    "permanent_address": "permanent_address",
    "city": "city",
    "state": "state",
    "country": "country",
    "pincode": "pincode",
    "address": "address",
    "pan number": "pan_number",
    "pan_number": "pan_number",
    "pan": "pan_number",
    "uan": "uan_number",
    "uan number": "uan_number",
    "uan_number": "uan_number",
    "bank account number": "bank_account",
    "bank account": "bank_account",
    "bank_account": "bank_account",
    "bank_account_number": "bank_account",
    "ifsc code": "bank_ifsc",
    "ifsc_code": "bank_ifsc",
    "ifsc": "bank_ifsc",
}

_DATE_FIELDS = {"date_of_joining", "date_of_birth", "confirmation_date"}
_DECIMAL_FIELDS = {"basic_salary", "ctc", "hra"}
_ENUM_FIELDS = {
    "employment_type": {"full_time", "part_time", "contract", "intern", "probation"},
    "status": {"active", "inactive", "pending", "on_leave", "terminated", "resigned", "deactivated", "suspended", "locked", "archived", "password_reset_required"},
    "gender": {"male", "female", "other"},
}
_REQUIRED_FIELDS = {"first_name", "last_name", "email", "job_title", "date_of_joining"}


def _normalise_header(h: str) -> str:
    return _COLUMN_MAP.get(h.strip().lower(), h.strip().lower())


def _parse_date(val, row_num: int, field: str, errors: list) -> Optional[date]:
    if not val or str(val).strip() == "":
        return None
    try:
        if isinstance(val, date):
            return val
        if isinstance(val, datetime):
            return val.date()
        cleaned = str(val).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
        # Excel serial date number
        try:
            from datetime import timedelta as td
            serial = float(cleaned)
            if 1 <= serial <= 2958465:
                return datetime(1899, 12, 30) + td(days=int(serial))
        except (ValueError, TypeError):
            pass
        errors.append({"row": row_num, "employee_id": "", "email": "", "field": field, "error": f"Invalid date for {field}: {val}"})
    except Exception:
        errors.append({"row": row_num, "employee_id": "", "email": "", "field": field, "error": f"Invalid date for {field}: {val}"})
    return None


def _parse_decimal(val, row_num: int, field: str, errors: list):
    if val is None or str(val).strip() == "":
        return None
    try:
        return Decimal(str(val).replace(",", ""))
    except Exception:
        errors.append({"row": row_num, "employee_id": "", "email": "", "field": field, "error": f"Invalid number for {field}: {val}"})
    return None


def _parse_enum(val, field: str):
    if not val or str(val).strip() == "":
        return None
    cleaned = str(val).strip().lower().replace(" ", "_").replace("-", "_")
    allowed = _ENUM_FIELDS.get(field, set())
    if cleaned in allowed:
        return cleaned
    if not allowed:
        return val
    return None


def import_employees_from_file(
    db: Session,
    file_bytes: bytes,
    filename: str,
    organization_id: int,
    current_user_id: int,
) -> dict:
    result = {
        "total_rows": 0,
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "failed": 0,
        "departments_created": 0,
        "designations_created": 0,
        "errors": [],
    }

    try:
        rows = _parse_file(file_bytes, filename, result)
    except Exception as e:
        result["errors"].append({"row": 0, "employee_id": "", "email": "", "field": "file", "error": f"Failed to parse file: {e}"})
        return result

    if not rows:
        return result

    seen_emails = set()
    seen_import_ids = set()
    created_for_email = []

    # Use savepoints so individual row failures don't rollback valid rows
    for row_num, row in enumerate(rows, start=2):
        norm = {}
        for k, v in row.items():
            field = _normalise_header(k)
            norm[field] = v
        row_data = norm

        employee_id_val = str(row_data.get("employee_id", "")).strip() if row_data.get("employee_id") else ""
        email_val = str(row_data.get("email", "")).strip() if row_data.get("email") else ""

        # Validate required fields
        missing = [f for f in _REQUIRED_FIELDS if not row_data.get(f) or str(row_data.get(f, "")).strip() == ""]
        if missing:
            result["skipped"] += 1
            result["errors"].append({"row": row_num, "employee_id": employee_id_val, "email": email_val, "field": ", ".join(missing), "error": f"Missing required fields: {', '.join(missing)}"})
            continue

        # Validate email format
        email_val = str(row_data["email"]).strip()
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email_val):
            result["skipped"] += 1
            result["errors"].append({"row": row_num, "employee_id": employee_id_val, "email": email_val, "field": "email", "error": "Invalid email format"})
            continue

        # Check duplicate by email
        existing = db.query(Employee).filter(Employee.email == email_val).first()
        if existing and existing.organization_id != organization_id:
            result["skipped"] += 1
            result["errors"].append({"row": row_num, "employee_id": employee_id_val, "email": email_val, "field": "email", "error": "Email already used in another organization"})
            continue

        # Check in-batch duplicates
        if email_val in seen_emails:
            result["skipped"] += 1
            result["errors"].append({"row": row_num, "employee_id": employee_id_val, "email": email_val, "field": "email", "error": "Duplicate email within import file"})
            continue
        seen_emails.add(email_val)

        if employee_id_val and employee_id_val in seen_import_ids:
            result["skipped"] += 1
            result["errors"].append({"row": row_num, "employee_id": employee_id_val, "email": email_val, "field": "employee_id", "error": "Duplicate employee ID within import file"})
            continue
        if employee_id_val:
            seen_import_ids.add(employee_id_val)

        # Parse fields
        payload = {
            "first_name": str(row_data.get("first_name", "")).strip(),
            "last_name": str(row_data.get("last_name", "")).strip(),
            "email": email_val,
            "phone": str(row_data.get("phone", "")).strip() or None,
            "job_title": str(row_data.get("job_title", "")).strip(),
            "work_email": str(row_data.get("work_email", "")).strip() or None,
            "personal_email": str(row_data.get("personal_email", "")).strip() or None,
            "company": str(row_data.get("company", "")).strip() or None,
            "business_unit": str(row_data.get("business_unit", "")).strip() or None,
            "division": str(row_data.get("division", "")).strip() or None,
            "team": str(row_data.get("team", "")).strip() or None,
            "current_address": str(row_data.get("current_address", "")).strip() or None,
            "permanent_address": str(row_data.get("permanent_address", "")).strip() or None,
            "city": str(row_data.get("city", "")).strip() or None,
            "state": str(row_data.get("state", "")).strip() or None,
            "country": str(row_data.get("country", "")).strip() or None,
            "pincode": str(row_data.get("pincode", "")).strip() or None,
            "address": str(row_data.get("address", "")).strip() or None,
        }

        # Date fields
        for f in _DATE_FIELDS:
            val = _parse_date(row_data.get(f), row_num, f, result["errors"])
            if val:
                payload[f] = val

        # Decimal fields
        for f in _DECIMAL_FIELDS:
            val = _parse_decimal(row_data.get(f), row_num, f, result["errors"])
            if val is not None:
                payload[f] = val

        # Employee profile fields (employee_profiles)
        profile_fields = {}
        for pf in ("pan_number", "uan_number", "bank_account", "bank_ifsc"):
            raw_val = row_data.get(pf)
            clean_val = str(raw_val).strip() if raw_val is not None else ""
            if clean_val:
                profile_fields[pf] = clean_val

        # Payroll roster fields (payroll_employees) — HRA is an ANNUAL amount
        payroll_fields = {}
        if "hra" in payload:
            payroll_fields["hra"] = payload.pop("hra")

        # Enum fields
        row_invalid = False
        for f in ("employment_type", "status", "gender"):
            raw = row_data.get(f)
            parsed = _parse_enum(raw, f)
            if parsed:
                payload[f] = parsed
            elif raw and str(raw).strip():
                result["skipped"] += 1
                result["errors"].append({"row": row_num, "employee_id": employee_id_val, "email": email_val, "field": f, "error": f"Invalid {f}: {raw}. Allowed: {', '.join(_ENUM_FIELDS.get(f, []))}"})
                row_invalid = True
                break
        if row_invalid:
            continue

        # Resolve department by name – auto-create if not found
        dept_name = str(row_data.get("department_name", "")).strip()
        if dept_name:
            dept = db.query(Department).filter(
                Department.name.ilike(dept_name),
                Department.organization_id == organization_id,
            ).first()
            if dept:
                payload["department_id"] = dept.id
            else:
                dept_code = "DEPT" + hashlib.md5(f"{dept_name}_{organization_id}".encode()).hexdigest()[:6].upper()
                dept = Department(
                    name=dept_name,
                    code=dept_code,
                    description="Auto-created from employee import",
                    organization_id=organization_id,
                )
                db.add(dept)
                db.flush()
                payload["department_id"] = dept.id
                result["departments_created"] += 1

        # Resolve designation by title – auto-create if not found
        designation_name = str(row_data.get("designation_name", "")).strip()
        if designation_name:
            desig = db.query(Designation).filter(
                Designation.title.ilike(designation_name),
                Designation.organization_id == organization_id,
            ).first()
            if desig:
                payload["designation_id"] = desig.id
            else:
                desig = Designation(
                    title=designation_name,
                    department_name=dept_name or None,
                    organization_id=organization_id,
                )
                db.add(desig)
                db.flush()
                payload["designation_id"] = desig.id
                result["designations_created"] += 1

        # Password
        password = str(row_data.get("password", "")).strip() if row_data.get("password") else None
        if not password:
            password = _generate_temp_password()

        # Use savepoint per row so failures don't rollback valid rows
        try:
            with db.begin_nested():
                if existing:
                    for field, value in payload.items():
                        if value is not None:
                            setattr(existing, field, value)
                    existing.updated_by = current_user_id
                    employee = existing
                    result["updated"] += 1
                else:
                    from app.core.code_generation import generate_employee_code
                    emp_data = {k: v for k, v in payload.items() if v is not None}
                    employee = Employee(
                        **emp_data,
                        hashed_password=hash_password(password),
                        employee_code=generate_employee_code(db, organization_id=organization_id),
                        organization_id=organization_id,
                        role=UserRole.EMPLOYEE,
                        is_active=True,
                        created_by=current_user_id,
                    )
                    db.add(employee)
                    result["created"] += 1

                # Persist imported profile fields (employee_profiles)
                if profile_fields:
                    profile = db.query(EmployeeProfile).filter(EmployeeProfile.employee_id == employee.id).first()
                    if profile:
                        for field, value in profile_fields.items():
                            setattr(profile, field, value)
                    else:
                        db.add(EmployeeProfile(
                            employee_id=employee.id,
                            organization_id=organization_id,
                            **profile_fields,
                        ))

                # Persist imported payroll roster fields (payroll_employees) — HRA is annual
                if payroll_fields:
                    payroll_emp = db.query(PayrollEmployee).filter(
                        PayrollEmployee.organization_id == organization_id,
                        PayrollEmployee.employee_code == employee.employee_code,
                    ).first()
                    if payroll_emp:
                        for field, value in payroll_fields.items():
                            setattr(payroll_emp, field, value)
                    else:
                        db.add(PayrollEmployee(
                            organization_id=organization_id,
                            employee_code=employee.employee_code,
                            first_name=payload.get("first_name") or "",
                            last_name=payload.get("last_name") or "",
                            email=email_val or None,
                            phone=payload.get("phone"),
                            department=dept_name or None,
                            designation=designation_name or None,
                            employment_type=payload.get("employment_type") or EmploymentType.FULL_TIME.value,
                            status=payload.get("status") or EmployeeStatus.ACTIVE.value,
                            date_of_joining=payload.get("date_of_joining"),
                            ctc=payload.get("ctc"),
                            **payroll_fields,
                        ))
            if not existing:
                created_for_email.append({
                    "email": email_val,
                    "employee_name": f"{payload.get('first_name') or ''} {payload.get('last_name') or ''}".strip(),
                    "temporary_password": password,
                })
        except Exception as e:
            result["failed"] += 1
            result["errors"].append({"row": row_num, "employee_id": employee_id_val, "email": email_val, "field": "general", "error": f"{'Update' if existing else 'Create'} failed: {str(e)[:200]}"})

        result["total_rows"] = row_num - 1

    import_committed = True
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        import_committed = False
        result["failed"] = result["total_rows"]
        result["created"] = 0
        result["updated"] = 0
        result["errors"].append({"row": 0, "employee_id": "", "email": "", "field": "general", "error": f"Bulk commit failed: {str(e)[:300]}"})

    if import_committed:
        for item in created_for_email:
            _notify_email(
                "send_employee_welcome_email",
                email=item["email"],
                employee_name=item["employee_name"] or item["email"],
                temporary_password=item["temporary_password"],
                organization_id=organization_id,
                db=db,
            )

    result["total_rows"] = len(rows)
    return result


def _parse_file(file_bytes: bytes, filename: str, result: dict) -> list[dict]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("xlsx", "xls"):
        return _parse_excel(file_bytes, result)
    elif ext == "csv":
        return _parse_csv(file_bytes, result)
    else:
        result["errors"].append({"row": 0, "employee_id": "", "email": "", "field": "file", "error": f"Unsupported file format: .{ext}. Use .xlsx, .xls, or .csv"})
        return []


def _parse_excel(file_bytes: bytes, result: dict) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        result["errors"].append({"row": 0, "employee_id": "", "email": "", "field": "file", "error": "Excel file has no sheets"})
        return []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(c).strip() if c else "" for c in next(rows_iter)]
    except StopIteration:
        result["errors"].append({"row": 0, "employee_id": "", "email": "", "field": "file", "error": "Excel file is empty"})
        return []

    parsed = []
    for row_idx, row in enumerate(rows_iter, start=2):
        record = {}
        has_data = False
        for col_idx, val in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                record[headers[col_idx]] = val
                if val is not None and str(val).strip():
                    has_data = True
        if has_data:
            parsed.append(record)

    wb.close()
    return parsed


def _parse_csv(file_bytes: bytes, result: dict) -> list[dict]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        result["errors"].append({"row": 0, "employee_id": "", "email": "", "field": "file", "error": "CSV file has no headers"})
        return []

    parsed = []
    for row_idx, row in enumerate(reader, start=2):
        cleaned = {k.strip(): v.strip() if v else "" for k, v in row.items()}
        has_data = any(v for v in cleaned.values())
        if has_data:
            parsed.append(cleaned)

    return parsed


def _generate_import_template_bytes() -> dict:
    """Generate sample import file bytes (.xlsx) and return as bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Import Template"

    headers = [
        "Employee ID", "First Name", "Last Name", "Email", "Password",
        "Phone", "Job Title", "Department", "Designation", "Reporting Manager",
        "Employment Type", "Status", "Date of Joining", "Date of Birth",
        "Gender", "Basic Salary", "HRA", "CTC", "Work Email", "Personal Email",
        "Company", "Division", "Team",
        "Current Address", "Permanent Address", "City", "State", "Country",
        "Pincode", "PAN Number", "UAN", "Bank Account Number", "IFSC Code",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample data row
    sample = [
        "ZO0001", "John", "Doe", "john.doe@example.com", "Pass@1234",
        "+91-9876543210", "Software Engineer", "Engineering", "Senior Developer", "Jane Smith",
        "Full Time", "Active", "2024-01-15", "1995-06-15",
        "Male", "75000", "180000", "1200000", "john@company.com", "john@gmail.com",
        "ZoikoOne", "Engineering", "Frontend",
        "123 Main St, Mumbai", "456 Oak Ave, Mumbai", "Mumbai", "Maharashtra", "India",
        "400001", "ABCDE1234F", "101234567890", "12345678901", "SBIN0001234",
    ]

    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()
    return buf.read()


def get_all_employees(
    db: Session,
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    department_id: Optional[int] = None,
    status: Optional[EmployeeStatus] = None,
    organization_id: Optional[int] = None,
    visible_roles: Optional[list] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(Employee)

    if organization_id:
        query = query.filter(Employee.organization_id == organization_id)

    if visible_roles:
        query = query.filter(Employee.role.in_(visible_roles))

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (Employee.first_name.ilike(search_term)) |
            (Employee.last_name.ilike(search_term)) |
            (Employee.email.ilike(search_term)) |
            (Employee.employee_id.ilike(search_term)) |
            (Employee.employee_code.ilike(search_term))
        )

    if department_id:
        query = query.filter(Employee.department_id == department_id)

    if status:
        query = query.filter(Employee.status == status)

    total = query.count()
    employees = query.offset((page - 1) * per_page).limit(per_page).all()

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    employees,
    }


def get_employees(
    db: Session,
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    department_id: Optional[int] = None,
    status: Optional[EmployeeStatus] = None,
    employment_type: Optional[EmploymentType] = None,
    organization_id: Optional[int] = None,
    visible_roles: Optional[list] = None,
) -> dict:
    per_page = min(per_page, 10000)
    query = db.query(Employee)

    if organization_id:
        query = query.filter(Employee.organization_id == organization_id)

    # Exclude administrative roles from employee listing if not explicitly provided
    if visible_roles:
        query = query.filter(Employee.role.in_(visible_roles))
    else:
        query = query.filter(Employee.role == UserRole.EMPLOYEE)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (Employee.first_name.ilike(search_term)) |
            (Employee.last_name.ilike(search_term)) |
            (Employee.email.ilike(search_term)) |
            (Employee.employee_id.ilike(search_term)) |
            (Employee.employee_code.ilike(search_term)) |
            (Employee.job_title.ilike(search_term))
        )

    if department_id:
        query = query.filter(Employee.department_id == department_id)

    if status:
        query = query.filter(Employee.status == status)

    if employment_type:
        query = query.filter(Employee.employment_type == employment_type)

    total = query.count()
    employees = query.offset((page - 1) * per_page).limit(per_page).all()

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    employees,
    }


def get_employee_by_id(db: Session, employee_id: int, organization_id: Optional[int] = None) -> Employee:
    query = db.query(Employee).filter(Employee.id == employee_id)
    if organization_id:
        query = query.filter(Employee.organization_id == organization_id)
    employee = query.first()
    if not employee:
        raise NotFoundException("Employee", employee_id)
    return employee


def update_employee(db: Session, employee_id: int, data: EmployeeUpdate, organization_id: Optional[int] = None) -> Employee:
    employee = get_employee_by_id(db, employee_id, organization_id)

    if data.department_id:
        dept_query = db.query(Department).filter(Department.id == data.department_id)
        if organization_id:
            dept_query = dept_query.filter(Department.organization_id == organization_id)
        dept = dept_query.first()
        if not dept:
            raise NotFoundException("Department", data.department_id)

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(employee, field, value)

    db.commit()
    db.refresh(employee)
    return employee


def deactivate_employee(db: Session, employee_id: int, organization_id: Optional[int] = None) -> Employee:
    employee = get_employee_by_id(db, employee_id, organization_id)
    employee.is_active = False
    employee.status = EmployeeStatus.TERMINATED

    event = EmployeeLifecycle(
        employee_id=employee_id,
        organization_id=employee.organization_id,
        event_type="exit",
        event_date=datetime.now().date(),
        status="completed",
        reason="Employee deactivated via admin action",
    )
    db.add(event)
    db.commit()
    db.refresh(employee)

    _notify_email(
        "send_employee_account_status_email",
        email=employee.email,
        employee_name=_full_name(employee),
        status="deactivated",
        organization_id=employee.organization_id,
        db=db,
    )

    return employee


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE DELETE (hybrid: active → deactivate, inactive → permanent delete)
# ═══════════════════════════════════════════════════════════════════════════════

def _soft_delete_employee_record(db: Session, employee_id: int) -> Employee:
    """Mark an employee as terminated without committing (for bulk operations)."""
    employee = get_employee_by_id(db, employee_id)
    employee.is_active = False
    employee.status = EmployeeStatus.TERMINATED
    event = EmployeeLifecycle(
        employee_id=employee_id,
        organization_id=employee.organization_id,
        event_type="exit",
        event_date=datetime.now().date(),
        status="completed",
        reason="Employee deactivated via admin action",
    )
    db.add(event)
    db.flush()
    return employee


# Hard-delete safety: FK columns that reference employees.id but must NOT be
# deleted (or nulled) when an employee is permanently removed. Rows matching a
# deleted employee in any of these tables block the deletion to protect
# platform audit / support history.
_EMPLOYEE_HARD_DELETE_PRESERVE = {
    ("super_admin_approval_history", "performed_by"),
    ("super_admin_support_tickets", "raised_by"),
}


def _hard_delete_employee(db: Session, employee_id: int) -> None:
    """Permanently remove an employee and every reference owned by them.

    Walks the SQLAlchemy metadata for every FK column pointing at employees.id:
      - NULLs nullable columns (audit/actor references are preserved),
      - DELETEs rows whose non-null FK belongs to the employee being removed,
      - BLOCKS the deletion if a protected audit reference still points at it.
    """
    preserved_hits = []
    for table in Base.metadata.tables.values():
        for column in table.columns:
            for fk in column.foreign_keys:
                if fk.column.table.name != "employees" or fk.column.name != "id":
                    continue
                key = (table.name, column.name)
                params = {"employee_id": employee_id}
                if key in _EMPLOYEE_HARD_DELETE_PRESERVE:
                    hit = db.execute(
                        text(f"SELECT 1 FROM {table.name} WHERE {column.name} = :employee_id LIMIT 1"),
                        params,
                    ).first()
                    if hit:
                        preserved_hits.append(key)
                elif column.nullable:
                    db.execute(
                        text(f"UPDATE {table.name} SET {column.name} = NULL WHERE {column.name} = :employee_id"),
                        params,
                    )
                else:
                    db.execute(
                        text(f"DELETE FROM {table.name} WHERE {column.name} = :employee_id"),
                        params,
                    )

    if preserved_hits:
        raise BadRequestException(
            f"Cannot permanently delete employee {employee_id}: protected audit reference(s) exist "
            f"({', '.join(f'{t}.{c}' for t, c in preserved_hits)})."
        )

    db.execute(text("DELETE FROM employees WHERE id = :employee_id"), {"employee_id": employee_id})


def delete_employee(
    db: Session,
    employee_id: int,
    organization_id: Optional[int] = None,
    current_user_id: Optional[int] = None,
) -> dict:
    """Hybrid delete for a single employee.

    Mirrors the existing UI behavior: the trash action deactivates active
    employees and permanently removes already-inactive ones.
    """
    employee = get_employee_by_id(db, employee_id)
    if employee.status == EmployeeStatus.ACTIVE:
        _soft_delete_employee_record(db, employee_id)
        db.commit()
        _notify_email(
            "send_employee_account_status_email",
            email=employee.email,
            employee_name=_full_name(employee),
            status="deactivated",
            organization_id=employee.organization_id,
            db=db,
        )
        return {
            "action": "deactivated",
            "message": f"Employee {employee_id} has been deactivated.",
        }
    _hard_delete_employee(db, employee_id)
    db.commit()
    return {
        "action": "deleted",
        "message": f"Employee {employee_id} has been permanently deleted.",
    }


def bulk_delete_employees(
    db: Session,
    employee_ids: List[int],
    organization_id: Optional[int] = None,
    current_user_id: Optional[int] = None,
) -> dict:
    """Delete multiple employees.

    Each row is processed in its own savepoint so a single failure does not
    roll back the rest. Active employees are deactivated; inactive ones are
    permanently removed.
    """
    result = {
        "deactivated": 0,
        "deleted": 0,
        "failed": 0,
        "total": len(employee_ids),
        "errors": [],
    }
    deactivated_for_email = []
    for employee_id in employee_ids:
        try:
            with db.begin_nested():
                employee = db.query(Employee).filter(Employee.id == employee_id).first()
                if not employee:
                    raise NotFoundException("Employee", employee_id)
                if organization_id and employee.organization_id != organization_id:
                    raise BadRequestException(
                        f"Access denied: employee {employee_id} does not belong to this organization"
                    )
                if employee.status == EmployeeStatus.ACTIVE:
                    _soft_delete_employee_record(db, employee_id)
                    result["deactivated"] += 1
                    deactivated_for_email.append(employee)
                else:
                    _hard_delete_employee(db, employee_id)
                    result["deleted"] += 1
        except Exception as exc:
            result["failed"] += 1
            result["errors"].append({
                "employee_id": employee_id,
                "error": str(exc)[:300],
            })
    db.commit()
    for employee in deactivated_for_email:
        _notify_email(
            "send_employee_account_status_email",
            email=employee.email,
            employee_name=_full_name(employee),
            status="deactivated",
            organization_id=employee.organization_id,
            db=db,
        )
    return result


def delete_all_employees(
    db: Session,
    organization_id: Optional[int] = None,
    current_user_id: Optional[int] = None,
) -> dict:
    """Deactivate every active employee and permanently remove every inactive
    employee in the organization (excluding the current admin, to avoid
    self-deletion)."""
    query = db.query(Employee).filter(Employee.role == UserRole.EMPLOYEE)
    if organization_id:
        query = query.filter(Employee.organization_id == organization_id)
    if current_user_id:
        query = query.filter(Employee.id != current_user_id)
    employee_ids = [row[0] for row in query.with_entities(Employee.id).all()]
    return bulk_delete_employees(db, employee_ids, organization_id=organization_id, current_user_id=current_user_id)


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

def get_employee_dashboard(db: Session, organization_id: Optional[int] = None) -> dict:
    base_filter = [Employee.organization_id == organization_id] if organization_id else []
    employee_filter = base_filter + [Employee.role == UserRole.EMPLOYEE]

    total = db.query(Employee).filter(*employee_filter).count()
    active = db.query(Employee).filter(*employee_filter, Employee.status == EmployeeStatus.ACTIVE).count()
    inactive = db.query(Employee).filter(*employee_filter, Employee.status != EmployeeStatus.ACTIVE).count()

    lc_filter = [EmployeeLifecycle.organization_id == organization_id] if organization_id else []
    probation = db.query(EmployeeLifecycle).filter(
        *lc_filter,
        EmployeeLifecycle.event_type == "probation_start",
        EmployeeLifecycle.status == "pending"
    ).count()

    from datetime import date as dt_date
    new_hires_this_month = db.query(Employee).filter(
        *employee_filter,
        extract("month", Employee.date_of_joining) == extract("month", dt_date.today()),
        extract("year", Employee.date_of_joining) == extract("year", dt_date.today())
    ).count()

    exits_this_month = db.query(Employee).filter(
        *employee_filter,
        extract("month", Employee.updated_at) == extract("month", dt_date.today()),
        extract("year", Employee.updated_at) == extract("year", dt_date.today()),
        Employee.status == EmployeeStatus.TERMINATED
    ).count()

    dept_breakdown = (
        db.query(Department.name, func.count(Employee.id))
        .join(Employee, Employee.department_id == Department.id, isouter=True)
        .filter(*employee_filter, Employee.status == EmployeeStatus.ACTIVE)
        .group_by(Department.name)
        .all()
    )

    designation_breakdown = (
        db.query(Employee.job_title, func.count(Employee.id))
        .filter(*employee_filter, Employee.status == EmployeeStatus.ACTIVE)
        .group_by(Employee.job_title)
        .all()
    )

    location_breakdown = (
        db.query(Employee.address, func.count(Employee.id))
        .filter(*employee_filter, Employee.status == EmployeeStatus.ACTIVE, Employee.address != None)
        .group_by(Employee.address)
        .all()
    )

    recent_lifecycle_events = []
    lifecycle_query = db.query(
        Employee.id, Employee.first_name, Employee.last_name,
        EmployeeLifecycle.event_type, EmployeeLifecycle.event_date,
        EmployeeLifecycle.status
    ).join(
        EmployeeLifecycle, Employee.id == EmployeeLifecycle.employee_id
    ).filter(*lc_filter).order_by(
        EmployeeLifecycle.created_at.desc()
    ).limit(10)

    for emp_id, first_name, last_name, event_type, event_date, status in lifecycle_query.all():
        recent_lifecycle_events.append({
            "employee_id": emp_id,
            "employee_name": f"{first_name} {last_name}",
            "event_type": event_type,
            "event_date": event_date,
            "status": status,
        })

    upcoming_probation_end = []
    for emp_id, first_name, last_name, event_date in db.query(
        Employee.id, Employee.first_name, Employee.last_name,
        EmployeeLifecycle.event_date
    ).join(
        EmployeeLifecycle, Employee.id == EmployeeLifecycle.employee_id
    ).filter(
        *lc_filter,
        EmployeeLifecycle.event_type == "probation_end",
        EmployeeLifecycle.status == "pending"
    ).order_by(EmployeeLifecycle.event_date).limit(5).all():
        upcoming_probation_end.append({
            "employee_id": emp_id,
            "employee_name": f"{first_name} {last_name}",
            "probation_end_date": event_date,
        })

    upcoming_confirmations = []
    for emp_id, first_name, last_name, event_date in db.query(
        Employee.id, Employee.first_name, Employee.last_name,
        EmployeeLifecycle.event_date
    ).join(
        EmployeeLifecycle, Employee.id == EmployeeLifecycle.employee_id
    ).filter(
        *lc_filter,
        EmployeeLifecycle.event_type == "confirmation",
        EmployeeLifecycle.status == "pending"
    ).order_by(EmployeeLifecycle.event_date).limit(5).all():
        upcoming_confirmations.append({
            "employee_id": emp_id,
            "employee_name": f"{first_name} {last_name}",
            "confirmation_date": event_date,
        })

    upcoming_anniversaries = []
    for emp_id, first_name, last_name, joining_date in db.query(
        Employee.id, Employee.first_name, Employee.last_name,
        Employee.date_of_joining
    ).filter(
        *base_filter,
        Employee.status == EmployeeStatus.ACTIVE,
        Employee.date_of_birth != None
    ).order_by(
        extract("month", Employee.date_of_birth),
        extract("day", Employee.date_of_birth)
    ).limit(5).all():
        today = dt_date.today()
        next_birthday = dt_date(today.year, joining_date.month, joining_date.day)
        if next_birthday < today:
            next_birthday = dt_date(today.year + 1, joining_date.month, joining_date.day)

        upcoming_anniversaries.append({
            "employee_id": emp_id,
            "employee_name": f"{first_name} {last_name}",
            "next_birthday": next_birthday,
            "join_date": joining_date,
        })

    return {
        "total_employees": total,
        "active_employees": active,
        "inactive_employees": inactive,
        "on_probation": probation,
        "new_hires_this_month": new_hires_this_month,
        "exits_this_month": exits_this_month,
        "department_distribution": [{"department": d, "count": c} for d, c in dept_breakdown],
        "designation_distribution": [{"designation": d, "count": c} for d, c in designation_breakdown],
        "location_distribution": [{"location": l, "count": c} for l, c in location_breakdown],
        "lifecycle_events": recent_lifecycle_events,
        "upcoming_probation_end": upcoming_probation_end,
        "upcoming_confirmations": upcoming_confirmations,
        "upcoming_anniversaries": upcoming_anniversaries,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE PROFILE
# ═══════════════════════════════════════════════════════════════════════════════

def get_employee_profile(db: Session, employee_id: int, **kwargs) -> EmployeeProfile:
    profile = db.query(EmployeeProfile).filter(EmployeeProfile.employee_id == employee_id).first()
    if not profile:
        raise NotFoundException("EmployeeProfile", employee_id)
    return profile


def create_employee_profile(db: Session, data) -> EmployeeProfile:
    profile = EmployeeProfile(**data.model_dump())
    db.add(profile)
    db.commit()
    db.refresh(profile)
    return profile


def update_employee_profile(db: Session, employee_id: int, data, organization_id: Optional[int] = None) -> EmployeeProfile:
    profile = db.query(EmployeeProfile).filter(EmployeeProfile.employee_id == employee_id).first()
    update_data = data.model_dump(exclude_unset=True)

    if not profile:
        profile = EmployeeProfile(
            employee_id=employee_id,
            organization_id=organization_id or 0,
            **update_data,
        )
        db.add(profile)
    else:
        for field, value in update_data.items():
            setattr(profile, field, value)

    db.commit()
    db.refresh(profile)
    return profile


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE REPORTING
# ═══════════════════════════════════════════════════════════════════════════════

def get_employee_reporting(db: Session, employee_id: int) -> EmployeeReporting:
    reporting = db.query(EmployeeReporting).filter(EmployeeReporting.employee_id == employee_id).first()
    if not reporting:
        raise NotFoundException("EmployeeReporting", employee_id)
    return reporting


def create_employee_reporting(db: Session, data) -> EmployeeReporting:
    reporting = EmployeeReporting(**data.model_dump())
    db.add(reporting)
    db.commit()
    db.refresh(reporting)
    return reporting


def update_employee_reporting(db: Session, employee_id: int, data) -> EmployeeReporting:
    reporting = get_employee_reporting(db, employee_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(reporting, field, value)
    db.commit()
    db.refresh(reporting)
    return reporting


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE LIFECYCLE
# ═══════════════════════════════════════════════════════════════════════════════

def get_employee_lifecycle(db: Session, employee_id: Optional[int] = None, organization_id: Optional[int] = None) -> list[EmployeeLifecycle]:
    query = db.query(EmployeeLifecycle)
    if organization_id:
        query = query.filter(EmployeeLifecycle.organization_id == organization_id)
    if employee_id:
        query = query.filter(EmployeeLifecycle.employee_id == employee_id)
    return query.order_by(EmployeeLifecycle.event_date.desc()).all()


def create_employee_lifecycle_event(db: Session, data) -> EmployeeLifecycle:
    event = EmployeeLifecycle(**data.model_dump())
    db.add(event)
    db.commit()
    db.refresh(event)

    return event


def update_employee_lifecycle_event(db: Session, event_id: int, data) -> EmployeeLifecycle:
    event = db.query(EmployeeLifecycle).filter(EmployeeLifecycle.id == event_id).first()
    if not event:
        raise NotFoundException("EmployeeLifecycle", event_id)

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(event, field, value)

    db.commit()
    db.refresh(event)
    return event


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE HISTORY
# ═══════════════════════════════════════════════════════════════════════════════

def get_employee_history(db: Session, employee_id: int) -> list[EmployeeHistory]:
    return db.query(EmployeeHistory).filter(
        EmployeeHistory.employee_id == employee_id
    ).order_by(EmployeeHistory.created_at.desc()).all()


def create_employee_history_entry(
    db: Session,
    employee_id: int,
    field_name: str,
    old_value: str,
    new_value: str,
    changed_by: Optional[int] = None,
    change_reason: Optional[str] = None,
) -> EmployeeHistory:
    history = EmployeeHistory(
        employee_id=employee_id,
        field_name=field_name,
        old_value=old_value,
        new_value=new_value,
        changed_by=changed_by,
        change_reason=change_reason,
    )
    db.add(history)
    db.commit()
    db.refresh(history)
    return history


# ═══════════════════════════════════════════════════════════════════════════════
# ORG CHART
# ═══════════════════════════════════════════════════════════════════════════════

def get_org_chart(db: Session, organization_id: int) -> dict:
    employees = db.query(
        Employee.id, Employee.first_name, Employee.last_name,
        Employee.job_title, Employee.department_id, Employee.status
    ).filter(
        Employee.organization_id == organization_id,
        Employee.status == EmployeeStatus.ACTIVE
    ).all()

    reporting = db.query(
        EmployeeReporting.employee_id, EmployeeReporting.manager_id
    ).filter(
        EmployeeReporting.organization_id == organization_id
    ).all()

    report_map = {r.employee_id: r.manager_id for r in reporting}

    departments = db.query(
        Department.id, Department.name
    ).filter(
        Department.id.in_([e.department_id for e in employees if e.department_id])
    ).all()

    dept_map = {d.id: d.name for d in departments}

    employee_map = {}
    for emp in employees:
        employee_map[emp.id] = {
            "id": emp.id,
            "name": f"{emp.first_name} {emp.last_name}",
            "job_title": emp.job_title,
            "department": dept_map.get(emp.department_id) if emp.department_id else None,
            "manager_id": report_map.get(emp.id),
            "status": emp.status,
            "children": [],
        }

    reporting_structure = []
    for emp in employees:
        manager_id = report_map.get(emp.id)
        if manager_id and manager_id in employee_map:
            employee_map[emp.id]["manager_name"] = employee_map[manager_id]["name"]
            employee_map[manager_id]["children"].append(employee_map[emp.id])
        else:
            reporting_structure.append(employee_map[emp.id])

    return {
        "employees": list(employee_map.values()),
        "reporting_structure": reporting_structure,
        "departments": dept_map,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE LIFECYCLE OPERATIONS
# ═══════════════════════════════════════════════════════════════════════════════

def change_manager(db: Session, data: ChangeManagerRequest) -> Employee:
    employee = get_employee_by_id(db, data.employee_id)

    reporting = db.query(EmployeeReporting).filter(
        EmployeeReporting.employee_id == data.employee_id
    ).first()

    old_manager_id = reporting.manager_id if reporting else None

    if not reporting:
        reporting = EmployeeReporting(
            employee_id=data.employee_id,
            organization_id=employee.organization_id or 1,
            manager_id=data.new_manager_id,
            effective_from=date.today(),
        )
        db.add(reporting)
    else:
        reporting.manager_id = data.new_manager_id

    db.commit()

    create_employee_history_entry(
        db, data.employee_id, "manager_id",
        str(old_manager_id), str(data.new_manager_id),
        change_reason=data.reason,
    )

    return employee


def confirm_probation(db: Session, data: ConfirmProbationRequest, organization_id: Optional[int] = None) -> EmployeeLifecycle:
    employee = get_employee_by_id(db, data.employee_id)

    employee.status = EmployeeStatus.ACTIVE
    employee.confirmation_date = data.confirmation_date

    event = EmployeeLifecycle(
        employee_id=data.employee_id,
        organization_id=employee.organization_id,
        event_type="confirmation",
        event_date=data.confirmation_date,
        status="completed",
        reason=data.notes,
    )
    db.add(event)
    db.commit()
    db.refresh(event)

    _notify_email(
        "send_employee_lifecycle_email",
        email=employee.email,
        employee_name=_full_name(employee),
        event_type="confirmation",
        effective_date=str(data.confirmation_date) if data.confirmation_date else "",
        details=data.notes or "",
        organization_id=employee.organization_id,
        db=db,
    )

    return event


def promote_employee(db: Session, data: PromoteEmployeeRequest, organization_id: Optional[int] = None) -> EmployeeLifecycle:
    employee = get_employee_by_id(db, data.employee_id)

    if data.new_designation_id:
        employee.designation_id = data.new_designation_id
    if data.new_salary:
        employee.basic_salary = data.new_salary

    event = EmployeeLifecycle(
        employee_id=data.employee_id,
        organization_id=employee.organization_id,
        event_type="promotion",
        event_date=data.effective_date,
        status="completed",
        new_value={"designation_id": data.new_designation_id, "salary": str(data.new_salary)},
        reason=data.reason,
    )
    db.add(event)
    db.commit()
    db.refresh(event)

    _notify_email(
        "send_employee_lifecycle_email",
        email=employee.email,
        employee_name=_full_name(employee),
        event_type="promotion",
        effective_date=str(data.effective_date) if data.effective_date else "",
        details=data.reason or "",
        organization_id=employee.organization_id,
        db=db,
    )

    return event


def transfer_employee(db: Session, data: TransferEmployeeRequest, organization_id: Optional[int] = None) -> EmployeeLifecycle:
    employee = get_employee_by_id(db, data.employee_id)

    if data.new_department_id:
        employee.department_id = data.new_department_id
    if data.new_manager_id:
        employee.reporting_manager_id = data.new_manager_id

    event = EmployeeLifecycle(
        employee_id=data.employee_id,
        organization_id=employee.organization_id,
        event_type="transfer",
        event_date=data.effective_date,
        status="completed",
        new_value={
            "department_id": data.new_department_id,
            "manager_id": data.new_manager_id,
            "location": data.new_location,
        },
        reason=data.reason,
    )
    db.add(event)
    db.commit()
    db.refresh(event)

    _notify_email(
        "send_employee_lifecycle_email",
        email=employee.email,
        employee_name=_full_name(employee),
        event_type="transfer",
        effective_date=str(data.effective_date) if data.effective_date else "",
        details=data.reason or "",
        organization_id=employee.organization_id,
        db=db,
    )

    return event


def resign_employee(db: Session, data: ResignationRequest, organization_id: Optional[int] = None) -> EmployeeLifecycle:
    employee = get_employee_by_id(db, data.employee_id)

    employee.status = EmployeeStatus.RESIGNED
    employee.is_active = False

    event = EmployeeLifecycle(
        employee_id=data.employee_id,
        organization_id=employee.organization_id,
        event_type="resignation",
        event_date=data.resignation_date,
        status="completed",
        new_value={
            "status": "resigned",
            "last_working_date": str(data.last_working_date),
        },
        reason=data.reason,
    )
    db.add(event)
    db.commit()
    db.refresh(event)

    _notify_email(
        "send_employee_lifecycle_email",
        email=employee.email,
        employee_name=_full_name(employee),
        event_type="resignation",
        effective_date=str(data.last_working_date) if data.last_working_date else "",
        details=data.reason or "",
        organization_id=employee.organization_id,
        db=db,
    )

    return event


def exit_employee(db: Session, data: ExitEmployeeRequest, organization_id: Optional[int] = None) -> EmployeeLifecycle:
    employee = get_employee_by_id(db, data.employee_id)

    employee.status = EmployeeStatus.TERMINATED
    employee.is_active = False

    event = EmployeeLifecycle(
        employee_id=data.employee_id,
        organization_id=employee.organization_id,
        event_type="exit",
        event_date=data.exit_date,
        status="completed",
        new_value={
            "status": data.exit_type,
            "final_settlement_date": str(data.final_settlement_date),
        },
        reason=data.reason,
    )
    db.add(event)
    db.commit()
    db.refresh(event)

    _notify_email(
        "send_employee_lifecycle_email",
        email=employee.email,
        employee_name=_full_name(employee),
        event_type="exit",
        effective_date=str(data.exit_date) if data.exit_date else "",
        details=data.reason or "",
        organization_id=employee.organization_id,
        db=db,
    )

    return event


def get_employee_reports(db: Session, filters: Optional[dict] = None, organization_id: Optional[int] = None) -> list:
    query = db.query(Employee)
    if organization_id:
        query = query.filter(Employee.organization_id == organization_id)
    if filters:
        if "department_id" in filters:
            query = query.filter(Employee.department_id == filters["department_id"])
        if "status" in filters:
            query = query.filter(Employee.status == filters["status"])
        if "search" in filters:
            search_term = f"%{filters['search']}%"
            query = query.filter(
                (Employee.first_name.ilike(search_term)) |
                (Employee.last_name.ilike(search_term)) |
                (Employee.email.ilike(search_term)) |
                (Employee.employee_id.ilike(search_term)) |
                (Employee.employee_code.ilike(search_term))
            )

    return query.order_by(Employee.created_at.desc()).all()


def export_employee_reports(db: Session, data: EmployeeExportRequest, organization_id: Optional[int] = None) -> list:
    return get_employee_reports(db, data.filters, organization_id)


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE COMPENSATION & BENEFITS
# ═══════════════════════════════════════════════════════════════════════════════

def create_employee_compensation(db: Session, data: EmployeeCompensationCreate, org_id: int) -> EmployeeCompensation:
    comp = EmployeeCompensation(**data.model_dump(), organization_id=org_id)
    db.add(comp)
    db.commit()
    db.refresh(comp)
    return comp

def get_employee_compensations(db: Session, org_id: int, employee_id: Optional[int] = None) -> list[EmployeeCompensation]:
    query = db.query(EmployeeCompensation).filter(EmployeeCompensation.organization_id == org_id)
    if employee_id:
        query = query.filter(EmployeeCompensation.employee_id == employee_id)
    return query.all()

def get_employee_compensation(db: Session, comp_id: int, org_id: int) -> EmployeeCompensation:
    comp = db.query(EmployeeCompensation).filter(EmployeeCompensation.id == comp_id, EmployeeCompensation.organization_id == org_id).first()
    if not comp:
        raise NotFoundException("EmployeeCompensation", comp_id)
    return comp

def update_employee_compensation(db: Session, comp_id: int, data: EmployeeCompensationUpdate, org_id: int) -> EmployeeCompensation:
    comp = db.query(EmployeeCompensation).filter(EmployeeCompensation.id == comp_id, EmployeeCompensation.organization_id == org_id).first()
    if not comp:
        raise NotFoundException("EmployeeCompensation", comp_id)
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(comp, key, value)
    db.commit()
    db.refresh(comp)
    return comp

def delete_employee_compensation(db: Session, comp_id: int, org_id: int) -> None:
    comp = db.query(EmployeeCompensation).filter(EmployeeCompensation.id == comp_id, EmployeeCompensation.organization_id == org_id).first()
    if not comp:
        raise NotFoundException("EmployeeCompensation", comp_id)
    from app.modules.hr.models import SalaryRevision
    db.query(SalaryRevision).filter(SalaryRevision.employee_compensation_id == comp_id).delete()
    db.delete(comp)
    db.commit()

def create_employee_benefit(db: Session, data: EmployeeBenefitCreate, org_id: int) -> EmployeeBenefit:
    emp_benefit = EmployeeBenefit(**data.model_dump(), organization_id=org_id)
    db.add(emp_benefit)
    db.commit()
    db.refresh(emp_benefit)
    return emp_benefit

def get_employee_benefits(db: Session, org_id: int) -> list[EmployeeBenefit]:
    return db.query(EmployeeBenefit).filter(EmployeeBenefit.organization_id == org_id).all()

def delete_employee_benefit(db: Session, emp_benefit_id: int, org_id: int) -> None:
    emp_benefit = db.query(EmployeeBenefit).filter(EmployeeBenefit.id == emp_benefit_id, EmployeeBenefit.organization_id == org_id).first()
    if not emp_benefit:
        raise NotFoundException("EmployeeBenefit", emp_benefit_id)
    db.delete(emp_benefit)
    db.commit()


def bulk_hard_delete_employees(db: Session, employee_ids: list[int], organization_id: int = None) -> dict:
    deleted = []
    failed = []
    for eid in employee_ids:
        try:
            q = db.query(Employee).filter(Employee.id == eid)
            if organization_id is not None:
                q = q.filter(Employee.organization_id == organization_id)
            emp = q.first()
            if not emp:
                failed.append({"id": eid, "reason": "Not found"})
                continue
            hard_delete_employee(db, eid, organization_id)
            deleted.append(eid)
        except Exception as ex:
            db.rollback()
            failed.append({"id": eid, "reason": str(ex)})
    return {"deleted": deleted, "failed": failed}


def hard_delete_employee(db: Session, employee_id: int, organization_id: int = None) -> None:
    q = db.query(Employee).filter(Employee.id == employee_id)
    if organization_id is not None:
        q = q.filter(Employee.organization_id == organization_id)
    employee = q.first()
    if not employee:
        raise NotFoundException("Employee", employee_id)

    # Reassign reportees to NULL
    rpt_q = db.query(Employee).filter(Employee.reporting_manager_id == employee_id)
    if organization_id is not None:
        rpt_q = rpt_q.filter(Employee.organization_id == organization_id)
    rpt_q.update({"reporting_manager_id": None}, synchronize_session=False)

    # Nullify self-referencing creator/updater FKs
    db.query(Employee).filter(
        Employee.created_by == employee_id,
    ).update({"created_by": None}, synchronize_session=False)
    db.query(Employee).filter(
        Employee.updated_by == employee_id,
    ).update({"updated_by": None}, synchronize_session=False)

    # Delete one-to-one profile/relationship records
    db.query(EmployeeProfile).filter(EmployeeProfile.employee_id == employee_id).delete(synchronize_session=False)
    db.query(EmployeeReporting).filter(EmployeeReporting.employee_id == employee_id).delete(synchronize_session=False)
    db.query(EmployeeLifecycle).filter(EmployeeLifecycle.employee_id == employee_id).delete(synchronize_session=False)
    db.query(EmployeeHistory).filter(EmployeeHistory.employee_id == employee_id).delete(synchronize_session=False)

    # Nullify other employees' reporting records that reference this employee as manager
    db.query(EmployeeReporting).filter(
        EmployeeReporting.manager_id == employee_id,
    ).update({"manager_id": None}, synchronize_session=False)
    db.query(EmployeeReporting).filter(
        EmployeeReporting.dotted_manager_id == employee_id,
    ).update({"dotted_manager_id": None}, synchronize_session=False)

    # Delete salary revisions before compensation records
    # FK: salary_revisions.employee_compensation_id -> employee_compensations.id
    emp_comp_ids = [r[0] for r in db.query(EmployeeCompensation.id).filter(
        EmployeeCompensation.employee_id == employee_id
    ).all()]
    if emp_comp_ids:
        db.query(SalaryRevision).filter(
            SalaryRevision.employee_compensation_id.in_(emp_comp_ids)
        ).delete(synchronize_session=False)

    # Delete compensation & allowance records
    db.query(EmployeeCompensation).filter(EmployeeCompensation.employee_id == employee_id).delete(synchronize_session=False)
    db.query(EmployeeBenefit).filter(EmployeeBenefit.employee_id == employee_id).delete(synchronize_session=False)
    db.query(Allowance).filter(Allowance.employee_id == employee_id).delete(synchronize_session=False)
    db.query(CompensationItem).filter(CompensationItem.employee_id == employee_id).delete(synchronize_session=False)

    # Delete HR operational records
    db.query(LeaveRequest).filter(LeaveRequest.employee_id == employee_id).delete(synchronize_session=False)
    db.query(LeaveBalance).filter(LeaveBalance.employee_id == employee_id).delete(synchronize_session=False)
    db.query(AttendanceRecord).filter(AttendanceRecord.employee_id == employee_id).delete(synchronize_session=False)
    db.query(ShiftRoster).filter(ShiftRoster.employee_id == employee_id).delete(synchronize_session=False)
    db.query(EssRequest).filter(EssRequest.employee_id == employee_id).delete(synchronize_session=False)

    # Delete performance records
    db.query(PerformanceGoal).filter(PerformanceGoal.employee_id == employee_id).delete(synchronize_session=False)
    db.query(PerformanceKpi).filter(PerformanceKpi.employee_id == employee_id).delete(synchronize_session=False)
    db.query(PerformanceFeedback).filter(PerformanceFeedback.employee_id == employee_id).delete(synchronize_session=False)
    db.query(Appraisal).filter(Appraisal.employee_id == employee_id).delete(synchronize_session=False)
    db.query(PerformanceReview).filter(PerformanceReview.employee_id == employee_id).delete(synchronize_session=False)

    # Delete learning records
    db.query(LearningEnrollment).filter(LearningEnrollment.employee_id == employee_id).delete(synchronize_session=False)
    db.query(LearningCertification).filter(LearningCertification.employee_id == employee_id).delete(synchronize_session=False)
    db.query(LearningSkill).filter(LearningSkill.employee_id == employee_id).delete(synchronize_session=False)
    db.query(LearningQuizAttempt).filter(LearningQuizAttempt.employee_id == employee_id).delete(synchronize_session=False)
    db.query(LearningTrainingProgramAssignment).filter(LearningTrainingProgramAssignment.employee_id == employee_id).delete(synchronize_session=False)

    # Delete recruitment records where employee was interviewer or approver
    db.query(RecruitmentInterviewFeedback).filter(
        RecruitmentInterviewFeedback.interviewer_id == employee_id
    ).delete(synchronize_session=False)
    db.query(RecruitmentOfferApproval).filter(
        RecruitmentOfferApproval.approver_id == employee_id
    ).delete(synchronize_session=False)

    # Delete travel approvals linked to employee's requests and where employee is approver
    emp_request_ids = [r[0] for r in db.query(TravelRequest.id).filter(TravelRequest.employee_id == employee_id).all()]
    if emp_request_ids:
        db.query(TravelApproval).filter(TravelApproval.request_id.in_(emp_request_ids)).delete(synchronize_session=False)
    db.query(TravelApproval).filter(TravelApproval.approver_id == employee_id).delete(synchronize_session=False)

    # Delete travel receipts linked to employee's expenses
    emp_expense_ids = [r[0] for r in db.query(TravelExpense.id).filter(TravelExpense.employee_id == employee_id).all()]
    if emp_expense_ids:
        db.query(TravelReceipt).filter(TravelReceipt.expense_id.in_(emp_expense_ids)).delete(synchronize_session=False)

    # Delete travel expense & request records (expenses first, FK: travel_receipts.expense_id -> travel_expenses.id)
    db.query(TravelExpense).filter(TravelExpense.employee_id == employee_id).delete(synchronize_session=False)
    db.query(TravelRequest).filter(TravelRequest.employee_id == employee_id).delete(synchronize_session=False)
    db.query(WfSuccession).filter(WfSuccession.employee_id == employee_id).delete(synchronize_session=False)

    # Delete time & compliance records
    db.query(TimeEntry).filter(TimeEntry.employee_id == employee_id).delete(synchronize_session=False)
    db.query(PolicyAcknowledgement).filter(PolicyAcknowledgement.employee_id == employee_id).delete(synchronize_session=False)
    db.query(ComplianceRecord).filter(ComplianceRecord.employee_id == employee_id).delete(synchronize_session=False)

    # Nullify employee_id on records where it's nullable (prevents FK violations for other employees' records)
    db.query(Asset).filter(Asset.employee_id == employee_id).update({"employee_id": None}, synchronize_session=False)
    db.query(HrDocument).filter(HrDocument.employee_id == employee_id).update({"employee_id": None}, synchronize_session=False)
    db.query(HrDocument).filter(HrDocument.approved_by == employee_id).update({"approved_by": None}, synchronize_session=False)
    db.query(HrDocument).filter(HrDocument.uploaded_by == employee_id).update({"uploaded_by": None}, synchronize_session=False)
    db.query(DocumentAssignment).filter(DocumentAssignment.employee_id == employee_id).delete(synchronize_session=False)
    db.query(DocumentAssignment).filter(DocumentAssignment.assigned_by == employee_id).update({"assigned_by": None}, synchronize_session=False)
    db.query(OnboardingNewHire).filter(OnboardingNewHire.employee_id == employee_id).update({"employee_id": None}, synchronize_session=False)
    db.query(OnboardingNewHire).filter(OnboardingNewHire.manager_id == employee_id).update({"manager_id": None}, synchronize_session=False)
    db.query(OnboardingPreboardingTask).filter(OnboardingPreboardingTask.employee_id == employee_id).update({"employee_id": None}, synchronize_session=False)

    # Delete records that are intrinsically owned by this employee (NOT NULL FK — can't nullify)
    db.query(EngagementSurvey).filter(EngagementSurvey.employee_id == employee_id).delete(synchronize_session=False)
    db.query(SupportTicket).filter(SupportTicket.raised_by == employee_id).delete(synchronize_session=False)
    db.query(ApprovalHistory).filter(ApprovalHistory.performed_by == employee_id).delete(synchronize_session=False)

    # Nullify remaining nullable FKs to this employee across HR, assets, org, learning & workforce planning
    db.query(Organization).filter(Organization.approved_by == employee_id).update({"approved_by": None}, synchronize_session=False)
    db.query(Holiday).filter(Holiday.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(LeaveRequest).filter(LeaveRequest.reviewed_by == employee_id).update({"reviewed_by": None}, synchronize_session=False)
    db.query(ShiftRoster).filter(ShiftRoster.assigned_by == employee_id).update({"assigned_by": None}, synchronize_session=False)
    db.query(AssetMaintenanceRequest).filter(AssetMaintenanceRequest.reported_by_id == employee_id).update({"reported_by_id": None}, synchronize_session=False)
    db.query(AssetMaintenanceRequest).filter(AssetMaintenanceRequest.resolved_by == employee_id).update({"resolved_by": None}, synchronize_session=False)
    db.query(AssetRequest).filter(AssetRequest.employee_id == employee_id).update({"employee_id": None}, synchronize_session=False)
    db.query(AssetRequest).filter(AssetRequest.approved_by == employee_id).update({"approved_by": None}, synchronize_session=False)
    db.query(AssetReport).filter(AssetReport.generated_by == employee_id).update({"generated_by": None}, synchronize_session=False)

    db.query(PerformanceReview).filter(PerformanceReview.reviewer_id == employee_id).update({"reviewer_id": None}, synchronize_session=False)
    db.query(PerformanceReview).filter(PerformanceReview.hr_reviewer_id == employee_id).update({"hr_reviewer_id": None}, synchronize_session=False)
    db.query(PerformanceReview).filter(PerformanceReview.admin_reviewer_id == employee_id).update({"admin_reviewer_id": None}, synchronize_session=False)
    db.query(PerformanceFeedback).filter(PerformanceFeedback.reviewer_id == employee_id).update({"reviewer_id": None}, synchronize_session=False)
    db.query(Appraisal).filter(Appraisal.reviewer_id == employee_id).update({"reviewer_id": None}, synchronize_session=False)
    db.query(Appraisal).filter(Appraisal.hr_reviewer_id == employee_id).update({"hr_reviewer_id": None}, synchronize_session=False)
    db.query(Appraisal).filter(Appraisal.admin_reviewer_id == employee_id).update({"admin_reviewer_id": None}, synchronize_session=False)

    db.query(RecruitmentInterview).filter(RecruitmentInterview.interviewer_id == employee_id).update({"interviewer_id": None}, synchronize_session=False)
    db.query(RecruitmentDocument).filter(RecruitmentDocument.uploaded_by == employee_id).update({"uploaded_by": None}, synchronize_session=False)
    db.query(TravelReceipt).filter(TravelReceipt.verified_by == employee_id).update({"verified_by": None}, synchronize_session=False)

    db.query(WfPlan).filter(WfPlan.owner_id == employee_id).update({"owner_id": None}, synchronize_session=False)
    db.query(WfPlan).filter(WfPlan.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(WfPlan).filter(WfPlan.updated_by == employee_id).update({"updated_by": None}, synchronize_session=False)
    db.query(WfHeadcount).filter(WfHeadcount.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(WfHeadcount).filter(WfHeadcount.updated_by == employee_id).update({"updated_by": None}, synchronize_session=False)
    db.query(WfSuccession).filter(WfSuccession.successor_employee_id == employee_id).update({"successor_employee_id": None}, synchronize_session=False)
    db.query(WfSuccession).filter(WfSuccession.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(WfSuccession).filter(WfSuccession.updated_by == employee_id).update({"updated_by": None}, synchronize_session=False)
    db.query(WfReport).filter(WfReport.generated_by == employee_id).update({"generated_by": None}, synchronize_session=False)

    db.query(LearningCourse).filter(LearningCourse.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(LearningPath).filter(LearningPath.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(LearningCertification).filter(LearningCertification.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(LearningAssessment).filter(LearningAssessment.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(LearningTrainingProgram).filter(LearningTrainingProgram.instructor_id == employee_id).update({"instructor_id": None}, synchronize_session=False)
    db.query(LearningTrainingProgram).filter(LearningTrainingProgram.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(LearningCalendarEvent).filter(LearningCalendarEvent.created_by == employee_id).update({"created_by": None}, synchronize_session=False)

    # Nullify FKs from OTHER employees' lifecycle/history/document records pointing at this employee
    db.query(EmployeeLifecycle).filter(EmployeeLifecycle.initiated_by == employee_id).update({"initiated_by": None}, synchronize_session=False)
    db.query(EmployeeLifecycle).filter(EmployeeLifecycle.approved_by == employee_id).update({"approved_by": None}, synchronize_session=False)
    db.query(EmployeeHistory).filter(EmployeeHistory.changed_by == employee_id).update({"changed_by": None}, synchronize_session=False)
    db.query(HrDocumentVersion).filter(HrDocumentVersion.uploaded_by == employee_id).update({"uploaded_by": None}, synchronize_session=False)
    db.query(DocumentApprovalStep).filter(DocumentApprovalStep.approved_by == employee_id).update({"approved_by": None}, synchronize_session=False)
    db.query(DocumentApprovalLog).filter(DocumentApprovalLog.performed_by == employee_id).update({"performed_by": None}, synchronize_session=False)
    db.query(CompliancePolicy).filter(CompliancePolicy.created_by == employee_id).update({"created_by": None}, synchronize_session=False)

    # Nullify FKs in reporting/insights, payroll & platform admin modules
    db.query(Report).filter(Report.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(ReportRun).filter(ReportRun.run_by == employee_id).update({"run_by": None}, synchronize_session=False)
    db.query(PayrollRun).filter(PayrollRun.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(PayrollRun).filter(PayrollRun.approved_by == employee_id).update({"approved_by": None}, synchronize_session=False)
    db.query(ComplianceDocument).filter(ComplianceDocument.uploaded_by == employee_id).update({"uploaded_by": None}, synchronize_session=False)
    db.query(PayrollActivityLog).filter(PayrollActivityLog.actor_id == employee_id).update({"actor_id": None}, synchronize_session=False)
    db.query(AuditLog).filter(AuditLog.performed_by == employee_id).update({"performed_by": None}, synchronize_session=False)
    db.query(Notification).filter(Notification.target_user_id == employee_id).update({"target_user_id": None}, synchronize_session=False)
    db.query(Notification).filter(Notification.created_by == employee_id).update({"created_by": None}, synchronize_session=False)
    db.query(SupportTicket).filter(SupportTicket.assigned_to == employee_id).update({"assigned_to": None}, synchronize_session=False)
    db.query(SecurityEvent).filter(SecurityEvent.user_id == employee_id).update({"user_id": None}, synchronize_session=False)
    db.query(SecurityEvent).filter(SecurityEvent.resolved_by == employee_id).update({"resolved_by": None}, synchronize_session=False)
    db.query(LoginActivity).filter(LoginActivity.user_id == employee_id).update({"user_id": None}, synchronize_session=False)

    # Nullify FKs across the billing module (created_by/updated_by/approved_by/etc. audit columns)
    for model, cols in [
        (BillingSetting, ("created_by", "updated_by")),
        (BillingCustomer, ("created_by", "updated_by")),
        (CustomerContact, ("created_by", "updated_by")),
        (ProductCategory, ("created_by", "updated_by")),
        (Product, ("created_by", "updated_by")),
        (PricingPlan, ("created_by", "updated_by")),
        (PriceList, ("created_by", "updated_by")),
        (PriceListItem, ("created_by", "updated_by")),
        (PricingRule, ("approved_by", "created_by", "updated_by")),
        (Discount, ("approved_by", "created_by", "updated_by")),
        (CurrencyPricing, ("created_by", "updated_by")),
        (TaxPricing, ("created_by", "updated_by")),
        (TaxGroup, ("created_by", "updated_by")),
        (Contract, ("created_by", "updated_by")),
        (ContractAmendment, ("changed_by",)),
        (Quotation, ("created_by", "updated_by")),
        (SubscriptionPlan, ("created_by", "updated_by")),
        (Subscription, ("created_by", "updated_by")),
        (SubscriptionEvent, ("created_by",)),
        (Invoice, ("created_by", "updated_by")),
        (InvoiceStatusHistory, ("changed_by",)),
        (PaymentMethod, ("created_by", "updated_by")),
        (Payment, ("created_by", "updated_by")),
        (PaymentAllocation, ("created_by",)),
        (CreditNote, ("created_by", "updated_by")),
        (CreditNoteApplication, ("created_by",)),
        (Refund, ("created_by", "updated_by")),
        (TaxRate, ("created_by", "updated_by")),
        (Tax, ("created_by",)),
        (DunningCase, ("created_by", "updated_by")),
        (CollectionsCase, ("assigned_to", "created_by", "updated_by")),
        (CollectionAction, ("performed_by",)),
        (RevenueRecognitionSchedule, ("created_by", "updated_by")),
        (BillingAuditLog, ("actor_id",)),
        (BillingConfiguration, ("created_by", "updated_by")),
        (CustomerDocument, ("uploaded_by",)),
        (CustomerNote, ("created_by", "updated_by")),
    ]:
        for col in cols:
            db.query(model).filter(getattr(model, col) == employee_id).update({col: None}, synchronize_session=False)

    db.flush()
    db.query(Employee).filter(Employee.id == employee_id).delete()
    db.commit()
