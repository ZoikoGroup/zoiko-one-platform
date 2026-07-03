import csv
import io
from datetime import datetime, timedelta
from typing import Optional
from sqlalchemy import func, asc, desc
from sqlalchemy.orm import Session

from app.modules.hr.models import (
    Asset, AssetMaintenanceRequest, AssetRequest, AssetCategory,
    AssetReport, AssetSetting, Employee,
    AssetCondition, AssetStatus, MaintenanceStatus, AssetRequestStatus, RequestPriority,
)
from app.modules.hr.schemas import (
    AssetCreate, AssetUpdate,
    MaintenanceCreate, MaintenanceUpdate, MaintenanceResolve,
    AssetRequestCreate,
    AssetCategoryCreate,
    AssetReportGenerate,
)
from app.core.exceptions import NotFoundException, AlreadyExistsException, BadRequestException
from app.core.sanitize import sanitize_input, sanitize_dict


def get_asset_dashboard(db: Session, organization_id: Optional[int] = None) -> dict:
    now = datetime.utcnow()
    thirty_days_ago = now - timedelta(days=30)

    base_filter = [Asset.deleted_at.is_(None)]
    cat_filter = [Asset.deleted_at.is_(None), Asset.category.isnot(None)]
    if organization_id:
        base_filter.append(Asset.organization_id == organization_id)
        cat_filter.append(Asset.organization_id == organization_id)

    total = db.query(func.count(Asset.id)).filter(*base_filter).scalar() or 0
    assigned = db.query(func.count(Asset.id)).filter(*base_filter, Asset.status == AssetStatus.ASSIGNED).scalar() or 0
    available = db.query(func.count(Asset.id)).filter(*base_filter, Asset.status == AssetStatus.AVAILABLE).scalar() or 0
    maintenance = db.query(func.count(Asset.id)).filter(*base_filter, Asset.status == AssetStatus.MAINTENANCE).scalar() or 0
    retired = db.query(func.count(Asset.id)).filter(*base_filter, Asset.status == AssetStatus.RETIRED).scalar() or 0
    lost = db.query(func.count(Asset.id)).filter(*base_filter, Asset.status == AssetStatus.LOST).scalar() or 0
    recently_added = db.query(func.count(Asset.id)).filter(*base_filter, Asset.created_at >= thirty_days_ago).scalar() or 0

    category_breakdown = (
        db.query(Asset.category, func.count(Asset.id))
        .filter(*cat_filter)
        .group_by(Asset.category)
        .all()
    )

    status_breakdown = (
        db.query(Asset.status, func.count(Asset.id))
        .filter(*base_filter)
        .group_by(Asset.status)
        .all()
    )

    req_filter = [AssetRequest.status == AssetRequestStatus.PENDING]
    maint_filter = [~AssetMaintenanceRequest.status.in_([MaintenanceStatus.RESOLVED, MaintenanceStatus.CANCELLED])]
    if organization_id:
        req_filter.append(AssetRequest.organization_id == organization_id)
        maint_filter.append(AssetMaintenanceRequest.organization_id == organization_id)

    pending_requests = db.query(func.count(AssetRequest.id)).filter(*req_filter).scalar() or 0
    open_maintenance = db.query(func.count(AssetMaintenanceRequest.id)).filter(*maint_filter).scalar() or 0

    return {
        "total_assets": total,
        "assigned_count": assigned,
        "available_count": available,
        "maintenance_count": maintenance,
        "retired_count": retired,
        "lost_count": lost,
        "recently_added": recently_added,
        "category_breakdown": [{"category": c, "count": cnt} for c, cnt in category_breakdown],
        "status_breakdown": [{"status": s.value, "count": cnt} for s, cnt in status_breakdown],
        "pending_requests": pending_requests,
        "open_maintenance": open_maintenance,
    }


def create_asset(db: Session, data: AssetCreate, organization_id: int = None) -> Asset:
    raw = data.model_dump()
    safe = sanitize_dict(raw)
    name = str(safe.get("name", ""))
    asset_tag = str(safe.get("asset_tag", ""))
    if not name.strip():
        raise BadRequestException("Asset name is required")
    if not asset_tag.strip():
        raise BadRequestException("Asset tag is required")
    safe["name"] = name
    safe["asset_tag"] = asset_tag

    existing = db.query(Asset.id).filter(
        Asset.asset_tag == asset_tag,
        Asset.deleted_at.is_(None),
        Asset.organization_id == organization_id,
    ).first()
    if existing:
        raise BadRequestException(f"Asset tag '{asset_tag}' already exists in this organization")

    # Ensure enum fields use Python enum instances, not plain strings
    enum_fields = {
        "condition": AssetCondition,
        "status": AssetStatus,
    }
    for fld, enum_cls in enum_fields.items():
        if fld in safe and isinstance(safe[fld], str):
            safe[fld] = enum_cls(safe[fld])

    asset = Asset(**safe)
    if organization_id:
        asset.organization_id = organization_id
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


SORTABLE_FIELDS = {
    "name": Asset.name,
    "asset_tag": Asset.asset_tag,
    "category": Asset.category,
    "department": Asset.department,
    "status": Asset.status,
    "condition": Asset.condition,
    "created_at": Asset.created_at,
    "updated_at": Asset.updated_at,
    "assigned_date": Asset.assigned_date,
    "purchase_date": Asset.purchase_date,
    "purchase_cost": Asset.purchase_cost,
}


def _get_asset_query(db: Session, search, status, category, department, employee_id, organization_id=None):
    query = db.query(Asset).filter(Asset.deleted_at.is_(None))

    if organization_id:
        query = query.filter(Asset.organization_id == organization_id)

    if search:
        search_term = f"%{search}%"
        query = query.outerjoin(Employee, Asset.employee_id == Employee.id).filter(
            (Asset.name.ilike(search_term)) |
            (Asset.asset_tag.ilike(search_term)) |
            (Asset.serial_number.ilike(search_term)) |
            (Asset.category.ilike(search_term)) |
            (Employee.first_name.ilike(search_term)) |
            (Employee.last_name.ilike(search_term))
        )

    if status:
        query = query.filter(Asset.status == status)

    if category:
        query = query.filter(Asset.category == category)

    if department:
        query = query.filter(Asset.department == department)

    if employee_id:
        query = query.filter(Asset.employee_id == employee_id)

    return query


def get_assets(
    db: Session,
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    status: Optional[AssetStatus] = None,
    category: Optional[str] = None,
    department: Optional[str] = None,
    employee_id: Optional[int] = None,
    sort_by: Optional[str] = "created_at",
    sort_order: Optional[str] = "desc",
    organization_id: Optional[int] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = _get_asset_query(db, search, status, category, department, employee_id, organization_id)

    total = query.count()

    sort_col = SORTABLE_FIELDS.get(sort_by, Asset.created_at)
    sort_fn = desc if sort_order == "desc" else asc
    assets = query.order_by(sort_fn(sort_col)).offset((page - 1) * per_page).limit(per_page).all()

    items = []
    for a in assets:
        items.append({
            "id": a.id,
            "employee_id": a.employee_id,
            "name": a.name,
            "asset_tag": a.asset_tag,
            "category": a.category,
            "serial_number": a.serial_number,
            "department": a.department,
            "assigned_date": a.assigned_date,
            "purchase_date": a.purchase_date,
            "purchase_cost": a.purchase_cost,
            "condition": a.condition,
            "status": a.status,
            "notes": a.notes,
            "created_at": a.created_at,
            "updated_at": a.updated_at,
            "employee_name": a.employee.full_name if a.employee else None,
        })

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "items": items,
    }


def get_asset_by_id(db: Session, asset_id: int, organization_id: Optional[int] = None) -> Asset:
    query = db.query(Asset).filter(Asset.id == asset_id, Asset.deleted_at.is_(None))
    if organization_id:
        query = query.filter(Asset.organization_id == organization_id)
    asset = query.first()
    if not asset:
        raise NotFoundException("Asset", asset_id)
    return asset


def update_asset(db: Session, asset_id: int, data: AssetUpdate, updated_by: int = None, organization_id: Optional[int] = None) -> Asset:
    asset = get_asset_by_id(db, asset_id, organization_id)
    update_data = sanitize_dict(data.model_dump(exclude_unset=True))
    if "name" in update_data and not str(update_data["name"]).strip():
        raise BadRequestException("Asset name cannot be empty")
    if "asset_tag" in update_data and not str(update_data["asset_tag"]).strip():
        raise BadRequestException("Asset tag cannot be empty")
    for field, value in update_data.items():
        setattr(asset, field, value)
    if updated_by is not None:
        asset.updated_by = updated_by
    db.commit()
    db.refresh(asset)
    return asset


def delete_asset(db: Session, asset_id: int, organization_id: Optional[int] = None) -> None:
    asset = get_asset_by_id(db, asset_id, organization_id)
    asset.deleted_at = datetime.utcnow()
    db.commit()


def assign_asset(db: Session, asset_id: int, employee_id: int, assigned_by: int, organization_id: Optional[int] = None) -> Asset:
    asset = get_asset_by_id(db, asset_id, organization_id)
    if asset.status == AssetStatus.RETIRED:
        raise BadRequestException("Cannot assign a retired asset")
    if asset.status == AssetStatus.LOST:
        raise BadRequestException("Cannot assign a lost asset")
    asset.employee_id = employee_id
    asset.status = AssetStatus.ASSIGNED
    asset.assigned_date = datetime.utcnow().date()
    asset.updated_by = assigned_by
    db.commit()
    db.refresh(asset)
    return asset


def return_asset(db: Session, asset_id: int, reason: str, condition: Optional[str] = None, returned_by: int = None, organization_id: Optional[int] = None) -> Asset:
    asset = get_asset_by_id(db, asset_id, organization_id)
    if asset.status != AssetStatus.ASSIGNED:
        raise BadRequestException("Asset is not currently assigned")
    asset.employee_id = None
    asset.status = AssetStatus.AVAILABLE
    asset.assigned_date = None
    if condition:
        asset.condition = condition
    if returned_by:
        asset.updated_by = returned_by
    db.commit()
    db.refresh(asset)
    return asset


def transfer_asset(db: Session, asset_id: int, new_employee_id: int, transferred_by: int, organization_id: Optional[int] = None) -> Asset:
    asset = get_asset_by_id(db, asset_id, organization_id)
    if asset.status == AssetStatus.RETIRED:
        raise BadRequestException("Cannot transfer a retired asset")
    if asset.status == AssetStatus.LOST:
        raise BadRequestException("Cannot transfer a lost asset")
    asset.employee_id = new_employee_id
    asset.status = AssetStatus.ASSIGNED
    asset.assigned_date = datetime.utcnow().date()
    asset.updated_by = transferred_by
    db.commit()
    db.refresh(asset)
    return asset


def export_assets_csv(db: Session, search=None, status=None, category=None, department=None, employee_id=None, sort_by="created_at", sort_order="desc", organization_id=None) -> str:
    query = _get_asset_query(db, search, status, category, department, employee_id, organization_id)
    sort_col = SORTABLE_FIELDS.get(sort_by, Asset.created_at)
    sort_fn = desc if sort_order == "desc" else asc
    assets = query.order_by(sort_fn(sort_col)).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Asset Tag", "Category", "Serial Number", "Department",
                     "Status", "Condition", "Assigned Date", "Purchase Date", "Purchase Cost",
                     "Employee Name", "Notes", "Created At", "Updated At"])
    for a in assets:
        writer.writerow([
            a.id, a.name, a.asset_tag, a.category, a.serial_number, a.department,
            a.status.value if a.status else "", a.condition.value if a.condition else "",
            a.assigned_date, a.purchase_date, a.purchase_cost,
            a.employee.full_name if a.employee else "", a.notes,
            a.created_at, a.updated_at,
        ])
    return output.getvalue()


def export_assets_excel(db: Session, search=None, status=None, category=None, department=None, employee_id=None, sort_by="created_at", sort_order="desc", organization_id=None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    query = _get_asset_query(db, search, status, category, department, employee_id, organization_id)
    sort_col = SORTABLE_FIELDS.get(sort_by, Asset.created_at)
    sort_fn = desc if sort_order == "desc" else asc
    assets = query.order_by(sort_fn(sort_col)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"

    headers = ["ID", "Name", "Asset Tag", "Category", "Serial Number", "Department",
               "Status", "Condition", "Assigned Date", "Purchase Date", "Purchase Cost",
               "Employee Name", "Notes", "Created At", "Updated At"]
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, a in enumerate(assets, 2):
        ws.cell(row=row_idx, column=1, value=a.id)
        ws.cell(row=row_idx, column=2, value=a.name)
        ws.cell(row=row_idx, column=3, value=a.asset_tag)
        ws.cell(row=row_idx, column=4, value=a.category)
        ws.cell(row=row_idx, column=5, value=a.serial_number)
        ws.cell(row=row_idx, column=6, value=a.department)
        ws.cell(row=row_idx, column=7, value=a.status.value if a.status else "")
        ws.cell(row=row_idx, column=8, value=a.condition.value if a.condition else "")
        ws.cell(row=row_idx, column=9, value=str(a.assigned_date) if a.assigned_date else "")
        ws.cell(row=row_idx, column=10, value=str(a.purchase_date) if a.purchase_date else "")
        ws.cell(row=row_idx, column=11, value=float(a.purchase_cost) if a.purchase_cost else "")
        ws.cell(row=row_idx, column=12, value=a.employee.full_name if a.employee else "")
        ws.cell(row=row_idx, column=13, value=a.notes)
        ws.cell(row=row_idx, column=14, value=str(a.created_at) if a.created_at else "")
        ws.cell(row=row_idx, column=15, value=str(a.updated_at) if a.updated_at else "")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def create_maintenance(db: Session, data: MaintenanceCreate, organization_id: int = None) -> AssetMaintenanceRequest:
    record = AssetMaintenanceRequest(**data.model_dump())
    if organization_id:
        record.organization_id = organization_id
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def get_maintenance_by_asset(db: Session, asset_id: int, organization_id: Optional[int] = None) -> list[AssetMaintenanceRequest]:
    query = db.query(AssetMaintenanceRequest).filter(
        AssetMaintenanceRequest.asset_id == asset_id
    )
    if organization_id:
        query = query.filter(AssetMaintenanceRequest.organization_id == organization_id)
    return query.order_by(AssetMaintenanceRequest.created_at.desc()).all()


def get_maintenance_by_id(db: Session, maintenance_id: int, organization_id: Optional[int] = None) -> AssetMaintenanceRequest:
    query = db.query(AssetMaintenanceRequest).filter(AssetMaintenanceRequest.id == maintenance_id)
    if organization_id:
        query = query.filter(AssetMaintenanceRequest.organization_id == organization_id)
    record = query.first()
    if not record:
        raise NotFoundException("AssetMaintenanceRequest", maintenance_id)
    return record


def update_maintenance(db: Session, maintenance_id: int, data: MaintenanceUpdate, organization_id: Optional[int] = None) -> AssetMaintenanceRequest:
    record = get_maintenance_by_id(db, maintenance_id, organization_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(record, field, value)
    db.commit()
    db.refresh(record)
    return record


def resolve_maintenance(db: Session, maintenance_id: int, data: MaintenanceResolve, organization_id: Optional[int] = None) -> AssetMaintenanceRequest:
    record = get_maintenance_by_id(db, maintenance_id, organization_id)
    record.status = MaintenanceStatus.RESOLVED
    record.resolution = data.resolution
    if data.resolved_by is not None:
        record.resolved_by = data.resolved_by
    record.resolved_on = datetime.utcnow()
    db.commit()
    db.refresh(record)
    return record


def create_asset_request(
    db: Session,
    data: AssetRequestCreate,
    organization_id: int = None,
    current_user=None,
) -> AssetRequest:
    req = AssetRequest(**data.model_dump(exclude={"employee_id", "employee_name"}))
    if organization_id:
        req.organization_id = organization_id
    if current_user:
        req.employee_id = current_user.id
        req.employee_name = current_user.full_name
    db.add(req)
    db.commit()
    db.refresh(req)
    return req


def get_asset_requests(
    db: Session,
    page: int = 1,
    per_page: int = 20,
    status: Optional[AssetRequestStatus] = None,
    priority: Optional[RequestPriority] = None,
    employee_id: Optional[int] = None,
    organization_id: Optional[int] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(AssetRequest)

    if organization_id:
        query = query.filter(AssetRequest.organization_id == organization_id)

    if status:
        query = query.filter(AssetRequest.status == status)

    if priority:
        query = query.filter(AssetRequest.priority == priority)

    if employee_id:
        query = query.filter(AssetRequest.employee_id == employee_id)

    total = query.count()
    requests = query.order_by(AssetRequest.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    items = []
    for req in requests:
        ename = req.employee_name
        if not ename and req.employee_id:
            emp = db.query(Employee).filter(Employee.id == req.employee_id).first()
            if emp:
                ename = f"{emp.first_name} {emp.last_name}"
        aname = None
        if req.approved_by:
            aemp = db.query(Employee).filter(Employee.id == req.approved_by).first()
            if aemp:
                aname = f"{aemp.first_name} {aemp.last_name}"
        items.append({
            "id": req.id,
            "employee_id": req.employee_id,
            "employee_name": ename,
            "asset_type": req.asset_type,
            "quantity": req.quantity,
            "priority": req.priority,
            "reason": req.reason,
            "notes": req.notes,
            "status": req.status,
            "requested_on": req.requested_on,
            "approved_by": req.approved_by,
            "approved_by_name": aname,
            "approved_on": req.approved_on,
            "fulfilled_on": req.fulfilled_on,
            "cancelled_on": req.cancelled_on,
            "created_at": req.created_at,
            "updated_at": req.updated_at,
        })

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "items": items,
    }


def _get_asset_request_by_id(db: Session, request_id: int, organization_id: Optional[int] = None) -> AssetRequest:
    query = db.query(AssetRequest).filter(AssetRequest.id == request_id)
    if organization_id:
        query = query.filter(AssetRequest.organization_id == organization_id)
    req = query.first()
    if not req:
        raise NotFoundException("AssetRequest", request_id)
    return req


def approve_asset_request(db: Session, request_id: int, approved_by: int, organization_id: Optional[int] = None) -> AssetRequest:
    req = _get_asset_request_by_id(db, request_id, organization_id)
    if req.status != AssetRequestStatus.PENDING:
        raise BadRequestException("Asset request is not in PENDING status")
    req.status = AssetRequestStatus.APPROVED
    req.approved_by = approved_by
    req.approved_on = datetime.utcnow()
    db.commit()
    db.refresh(req)
    return req


def reject_asset_request(db: Session, request_id: int, organization_id: Optional[int] = None) -> AssetRequest:
    req = _get_asset_request_by_id(db, request_id, organization_id)
    if req.status != AssetRequestStatus.PENDING:
        raise BadRequestException("Asset request is not in PENDING status")
    req.status = AssetRequestStatus.REJECTED
    db.commit()
    db.refresh(req)
    return req


def fulfill_asset_request(db: Session, request_id: int, organization_id: Optional[int] = None) -> AssetRequest:
    req = _get_asset_request_by_id(db, request_id, organization_id)
    if req.status != AssetRequestStatus.APPROVED:
        raise BadRequestException("Asset request is not in APPROVED status")
    req.status = AssetRequestStatus.FULFILLED
    req.fulfilled_on = datetime.utcnow()
    db.commit()
    db.refresh(req)
    return req


def cancel_asset_request(db: Session, request_id: int, organization_id: Optional[int] = None) -> AssetRequest:
    req = _get_asset_request_by_id(db, request_id, organization_id)
    if req.status not in (AssetRequestStatus.PENDING, AssetRequestStatus.APPROVED):
        raise BadRequestException("Asset request cannot be cancelled in its current status")
    req.status = AssetRequestStatus.CANCELLED
    req.cancelled_on = datetime.utcnow()
    db.commit()
    db.refresh(req)
    return req


def create_asset_category(db: Session, data: AssetCategoryCreate, organization_id: int = None) -> AssetCategory:
    query = db.query(AssetCategory).filter(AssetCategory.name.ilike(data.name))
    if organization_id:
        query = query.filter(AssetCategory.organization_id == organization_id)
    existing = query.first()
    if existing:
        raise AlreadyExistsException("AssetCategory", "name")
    category = AssetCategory(**data.model_dump())
    if organization_id:
        category.organization_id = organization_id
    db.add(category)
    db.commit()
    db.refresh(category)
    return category


def get_asset_categories(db: Session, organization_id: Optional[int] = None) -> list[AssetCategory]:
    query = db.query(AssetCategory).filter(AssetCategory.is_active == True)
    if organization_id:
        query = query.filter(AssetCategory.organization_id == organization_id)
    return query.order_by(AssetCategory.name).all()


def delete_asset_category(db: Session, category_id: int, organization_id: Optional[int] = None) -> None:
    query = db.query(AssetCategory).filter(AssetCategory.id == category_id)
    if organization_id:
        query = query.filter(AssetCategory.organization_id == organization_id)
    category = query.first()
    if not category:
        raise NotFoundException("AssetCategory", category_id)
    db.delete(category)
    db.commit()


def update_asset_category(db: Session, category_id: int, data: AssetCategoryCreate, organization_id: Optional[int] = None) -> AssetCategory:
    query = db.query(AssetCategory).filter(AssetCategory.id == category_id)
    if organization_id:
        query = query.filter(AssetCategory.organization_id == organization_id)
    category = query.first()
    if not category:
        raise NotFoundException("AssetCategory", category_id)
    existing = db.query(AssetCategory).filter(
        AssetCategory.name.ilike(data.name),
        AssetCategory.id != category_id,
    )
    if organization_id:
        existing = existing.filter(AssetCategory.organization_id == organization_id)
    if existing.first():
        raise AlreadyExistsException("AssetCategory", "name")
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(category, field, value)
    db.commit()
    db.refresh(category)
    return category


def create_asset_report(db: Session, data: AssetReportGenerate, generated_by: int) -> AssetReport:
    report = AssetReport(
        report_type=data.report_type,
        title=data.title,
        description=data.description,
        parameters=data.parameters,
        generated_by=generated_by,
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    return report


def get_asset_reports(db: Session) -> list[AssetReport]:
    return db.query(AssetReport).order_by(AssetReport.created_at.desc()).all()


def get_asset_settings(db: Session) -> list[AssetSetting]:
    return db.query(AssetSetting).order_by(AssetSetting.setting_key).all()


def update_asset_setting(db: Session, setting_key: str, setting_value: str, updated_by: int) -> AssetSetting:
    setting = db.query(AssetSetting).filter(AssetSetting.setting_key == setting_key).first()
    if setting:
        setting.setting_value = setting_value
        setting.updated_by = updated_by
    else:
        setting = AssetSetting(
            setting_key=setting_key,
            setting_value=setting_value,
            updated_by=updated_by,
        )
        db.add(setting)
    db.commit()
    db.refresh(setting)
    return setting
