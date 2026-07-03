from datetime import datetime, date
from typing import Optional
import json
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.modules.hr.models import (
    LearningCourse, LearningEnrollment, LearningPath, LearningPathItem,
    LearningCertification, LearningSkill, LearningAssessment,
    LearningAssessmentQuestion, LearningQuizAttempt,
    LearningTrainingProgram, LearningTrainingProgramAssignment,
    LearningCalendarEvent,
)
from app.modules.hr.schemas import (
    CourseCreate, CourseUpdate,
    EnrollmentCreate, EnrollmentUpdate,
    LearningPathCreate, LearningPathUpdate, LearningPathItemCreate, LearningPathItemResponse,
    CertificationCreate, CertificationUpdate,
    SkillCreate, SkillUpdate,
    AssessmentCreate, AssessmentUpdate,
    QuestionCreate, QuestionUpdate,
    QuizAttemptStart, QuizAttemptSubmit,
    TrainingProgramCreate, TrainingProgramUpdate,
    ProgramAssignmentCreate, ProgramAssignmentUpdate,
    CalendarEventCreate, CalendarEventUpdate,
)
from sqlalchemy.exc import SQLAlchemyError
from app.core.exceptions import NotFoundException, BadRequestException


def create_course(db: Session, data: CourseCreate, created_by: int = None, organization_id: Optional[int] = None) -> LearningCourse:
    course = LearningCourse(**data.model_dump(), created_by=created_by)
    if organization_id is not None:
        course.organization_id = organization_id
    db.add(course)
    db.commit()
    db.refresh(course)
    return course


def get_courses(
    db: Session,
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    course_type: Optional[str] = None,
    organization_id: Optional[int] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(LearningCourse)

    if search:
        term = f"%{search}%"
        query = query.filter(LearningCourse.course_name.ilike(term))

    if category:
        query = query.filter(LearningCourse.category == category)

    if status:
        query = query.filter(LearningCourse.status == status)

    if course_type:
        query = query.filter(LearningCourse.course_type == course_type)

    if organization_id is not None:
        query = query.filter(LearningCourse.organization_id == organization_id)

    total = query.count()
    courses = query.order_by(LearningCourse.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "items": courses,
    }


def get_course_by_id(db: Session, course_id: int, organization_id: Optional[int] = None) -> LearningCourse:
    course = db.query(LearningCourse).filter(LearningCourse.id == course_id).first()
    if not course:
        raise NotFoundException("LearningCourse", course_id)
    return course


def update_course(db: Session, course_id: int, data: CourseUpdate, organization_id: Optional[int] = None) -> LearningCourse:
    course = get_course_by_id(db, course_id, organization_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(course, field, value)
    db.commit()
    db.refresh(course)
    return course


def delete_course(db: Session, course_id: int, organization_id: Optional[int] = None) -> None:
    course = get_course_by_id(db, course_id, organization_id)
    db.delete(course)
    db.commit()


def create_enrollment(db: Session, data: EnrollmentCreate, organization_id: Optional[int] = None) -> LearningEnrollment:
    get_course_by_id(db, data.course_id)
    enrollment = LearningEnrollment(**data.model_dump())
    if organization_id is not None:
        enrollment.organization_id = organization_id
    db.add(enrollment)
    db.commit()
    db.refresh(enrollment)
    return enrollment


def get_enrollments(
    db: Session,
    page: int = 1,
    per_page: int = 20,
    employee_id: Optional[int] = None,
    course_id: Optional[int] = None,
    status: Optional[str] = None,
    organization_id: Optional[int] = None,
) -> dict:
    per_page = min(per_page, 100)
    query = db.query(LearningEnrollment)

    if employee_id:
        query = query.filter(LearningEnrollment.employee_id == employee_id)

    if course_id:
        query = query.filter(LearningEnrollment.course_id == course_id)

    if status:
        query = query.filter(LearningEnrollment.status == status)

    if organization_id is not None:
        query = query.filter(LearningEnrollment.organization_id == organization_id)

    total = query.count()
    enrollments = query.order_by(LearningEnrollment.enrolled_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    items = []
    for e in enrollments:
        items.append({
            "id": e.id,
            "course_id": e.course_id,
            "employee_id": e.employee_id,
            "status": e.status,
            "progress_pct": e.progress_pct,
            "enrolled_at": e.enrolled_at,
            "started_at": e.started_at,
            "completed_at": e.completed_at,
            "score": e.score,
            "notes": e.notes,
            "created_at": e.created_at,
            "updated_at": e.updated_at,
            "course_name": e.course.course_name if e.course else None,
            "employee_name": e.employee.full_name if e.employee else None,
        })

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "items": items,
    }


def get_enrollment_by_id(db: Session, enrollment_id: int) -> LearningEnrollment:
    enrollment = db.query(LearningEnrollment).filter(LearningEnrollment.id == enrollment_id).first()
    if not enrollment:
        raise NotFoundException("LearningEnrollment", enrollment_id)
    return enrollment


def update_enrollment(db: Session, enrollment_id: int, data: EnrollmentUpdate, organization_id: Optional[int] = None) -> LearningEnrollment:
    enrollment = get_enrollment_by_id(db, enrollment_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(enrollment, field, value)
    db.commit()
    db.refresh(enrollment)
    return enrollment


def delete_enrollment(db: Session, enrollment_id: int, organization_id: Optional[int] = None) -> None:
    enrollment = get_enrollment_by_id(db, enrollment_id)
    db.delete(enrollment)
    db.commit()


def create_learning_path(db: Session, data: LearningPathCreate, created_by: int, organization_id: Optional[int] = None) -> LearningPath:
    path = LearningPath(**data.model_dump(), created_by=created_by)
    if organization_id is not None:
        path.organization_id = organization_id
    db.add(path)
    db.commit()
    db.refresh(path)
    return path


def get_learning_paths(db: Session, organization_id: Optional[int] = None) -> list[dict]:
    query = db.query(LearningPath)
    if organization_id is not None:
        query = query.filter(LearningPath.organization_id == organization_id)
    paths = query.order_by(LearningPath.created_at.desc()).all()
    result = []
    for path in paths:
        items = []
        for item in path.items:
            items.append({
                "id": item.id,
                "path_id": item.path_id,
                "course_id": item.course_id,
                "sort_order": item.sort_order,
                "is_required": item.is_required,
                "course_name": item.course.course_name if item.course else None,
            })
        result.append({
            "id": path.id,
            "name": path.name,
            "description": path.description,
            "created_by": path.created_by,
            "is_active": path.is_active,
            "created_at": path.created_at,
            "updated_at": path.updated_at,
            "items": items,
        })
    return result


def get_learning_path_by_id(db: Session, path_id: int, organization_id: Optional[int] = None) -> LearningPath:
    path = db.query(LearningPath).filter(LearningPath.id == path_id).first()
    if not path:
        raise NotFoundException("LearningPath", path_id)
    return path


def update_learning_path(db: Session, path_id: int, data: LearningPathUpdate, organization_id: Optional[int] = None) -> LearningPath:
    path = get_learning_path_by_id(db, path_id, organization_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(path, field, value)
    db.commit()
    db.refresh(path)
    return path


def delete_learning_path(db: Session, path_id: int, organization_id: Optional[int] = None) -> None:
    path = get_learning_path_by_id(db, path_id, organization_id)
    db.delete(path)
    db.commit()


def add_path_item(db: Session, path_id: int, data: LearningPathItemCreate, organization_id: Optional[int] = None) -> LearningPathItem:
    get_learning_path_by_id(db, path_id, organization_id)
    get_course_by_id(db, data.course_id)
    item = LearningPathItem(**data.model_dump(), path_id=path_id)
    if organization_id is not None:
        item.organization_id = organization_id
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def remove_path_item(db: Session, path_id: int, item_id: int, organization_id: Optional[int] = None) -> None:
    get_learning_path_by_id(db, path_id, organization_id)
    item = db.query(LearningPathItem).filter(LearningPathItem.id == item_id, LearningPathItem.path_id == path_id).first()
    if not item:
        raise NotFoundException("LearningPathItem", item_id)
    db.delete(item)
    db.commit()


def update_path_item(db: Session, path_id: int, item_id: int, data: LearningPathItemCreate, organization_id: Optional[int] = None) -> LearningPathItem:
    get_learning_path_by_id(db, path_id, organization_id)
    item = db.query(LearningPathItem).filter(LearningPathItem.id == item_id, LearningPathItem.path_id == path_id).first()
    if not item:
        raise NotFoundException("LearningPathItem", item_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(item, field, value)
    db.commit()
    db.refresh(item)
    return item


def create_certification(db: Session, data: CertificationCreate, created_by: int, organization_id: Optional[int] = None) -> LearningCertification:
    cert = LearningCertification(**data.model_dump(exclude={'created_by'}), created_by=created_by)
    if organization_id is not None:
        cert.organization_id = organization_id
    db.add(cert)
    db.commit()
    db.refresh(cert)
    return cert


def get_certifications(db: Session, employee_id: Optional[int] = None, organization_id: Optional[int] = None) -> list[LearningCertification]:
    query = db.query(LearningCertification)
    if employee_id:
        query = query.filter(LearningCertification.employee_id == employee_id)
    if organization_id is not None:
        query = query.filter(LearningCertification.organization_id == organization_id)
    return query.order_by(LearningCertification.issue_date.desc()).all()


def get_certification_by_id(db: Session, cert_id: int) -> LearningCertification:
    cert = db.query(LearningCertification).filter(LearningCertification.id == cert_id).first()
    if not cert:
        raise NotFoundException("LearningCertification", cert_id)
    return cert


def update_certification(db: Session, cert_id: int, data: CertificationUpdate, organization_id: Optional[int] = None) -> LearningCertification:
    cert = get_certification_by_id(db, cert_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(cert, field, value)
    db.commit()
    db.refresh(cert)
    return cert


def delete_certification(db: Session, cert_id: int, organization_id: Optional[int] = None) -> None:
    cert = get_certification_by_id(db, cert_id)
    db.delete(cert)
    db.commit()


def create_skill(db: Session, data: SkillCreate, organization_id: Optional[int] = None) -> LearningSkill:
    skill = LearningSkill(**data.model_dump())
    if organization_id is not None:
        skill.organization_id = organization_id
    db.add(skill)
    db.commit()
    db.refresh(skill)
    return skill


def get_skills(db: Session, employee_id: Optional[int] = None) -> list[LearningSkill]:
    query = db.query(LearningSkill)
    if employee_id:
        query = query.filter(LearningSkill.employee_id == employee_id)
    return query.order_by(LearningSkill.skill_name).all()


def get_skill_by_id(db: Session, skill_id: int) -> LearningSkill:
    skill = db.query(LearningSkill).filter(LearningSkill.id == skill_id).first()
    if not skill:
        raise NotFoundException("LearningSkill", skill_id)
    return skill


def update_skill(db: Session, skill_id: int, data: SkillUpdate) -> LearningSkill:
    skill = get_skill_by_id(db, skill_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(skill, field, value)
    db.commit()
    db.refresh(skill)
    return skill


def delete_skill(db: Session, skill_id: int) -> None:
    skill = get_skill_by_id(db, skill_id)
    db.delete(skill)
    db.commit()



def create_assessment(db: Session, data: AssessmentCreate, created_by: int) -> LearningAssessment:
    get_course_by_id(db, data.course_id)
    assessment = LearningAssessment(**data.model_dump(), created_by=created_by)
    db.add(assessment)
    db.commit()
    db.refresh(assessment)
    return assessment


def get_assessments(db: Session, course_id: Optional[int] = None) -> list[LearningAssessment]:
    query = db.query(LearningAssessment)
    if course_id:
        query = query.filter(LearningAssessment.course_id == course_id)
    return query.order_by(LearningAssessment.created_at.desc()).all()


def get_assessment_by_id(db: Session, assessment_id: int) -> LearningAssessment:
    assessment = db.query(LearningAssessment).filter(LearningAssessment.id == assessment_id).first()
    if not assessment:
        raise NotFoundException("LearningAssessment", assessment_id)
    return assessment


def update_assessment(db: Session, assessment_id: int, data: AssessmentUpdate) -> LearningAssessment:
    assessment = get_assessment_by_id(db, assessment_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(assessment, field, value)
    db.commit()
    db.refresh(assessment)
    return assessment


def delete_assessment(db: Session, assessment_id: int) -> None:
    assessment = get_assessment_by_id(db, assessment_id)
    db.delete(assessment)
    db.commit()


def add_question(db: Session, assessment_id: int, data: QuestionCreate) -> LearningAssessmentQuestion:
    get_assessment_by_id(db, assessment_id)
    question = LearningAssessmentQuestion(**data.model_dump(), assessment_id=assessment_id)
    db.add(question)
    db.commit()
    db.refresh(question)
    return question


def get_questions(db: Session, assessment_id: int) -> list[LearningAssessmentQuestion]:
    return db.query(LearningAssessmentQuestion).filter(
        LearningAssessmentQuestion.assessment_id == assessment_id
    ).order_by(LearningAssessmentQuestion.sort_order).all()


def update_question(db: Session, question_id: int, data: QuestionUpdate) -> LearningAssessmentQuestion:
    question = db.query(LearningAssessmentQuestion).filter(LearningAssessmentQuestion.id == question_id).first()
    if not question:
        raise NotFoundException("LearningAssessmentQuestion", question_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(question, field, value)
    db.commit()
    db.refresh(question)
    return question


def delete_question(db: Session, question_id: int) -> None:
    question = db.query(LearningAssessmentQuestion).filter(LearningAssessmentQuestion.id == question_id).first()
    if not question:
        raise NotFoundException("LearningAssessmentQuestion", question_id)
    db.delete(question)
    db.commit()


def start_quiz(db: Session, data: QuizAttemptStart) -> LearningQuizAttempt:
    assessment = get_assessment_by_id(db, data.assessment_id)

    if not assessment.is_active:
        raise BadRequestException("Assessment is not active.")

    if assessment.max_attempts and assessment.max_attempts > 0:
        existing_count = db.query(LearningQuizAttempt).filter(
            LearningQuizAttempt.assessment_id == data.assessment_id,
            LearningQuizAttempt.employee_id == data.employee_id,
        ).count()
        if existing_count >= assessment.max_attempts:
            raise BadRequestException("Maximum number of attempts reached.")

    attempt_number = 1
    last_attempt = db.query(LearningQuizAttempt).filter(
        LearningQuizAttempt.assessment_id == data.assessment_id,
        LearningQuizAttempt.employee_id == data.employee_id,
    ).order_by(LearningQuizAttempt.attempt_number.desc()).first()
    if last_attempt:
        attempt_number = last_attempt.attempt_number + 1

    attempt = LearningQuizAttempt(
        assessment_id=data.assessment_id,
        employee_id=data.employee_id,
        enrollment_id=data.enrollment_id,
        started_at=datetime.utcnow(),
        attempt_number=attempt_number,
        status="in_progress",
    )
    db.add(attempt)
    db.commit()
    db.refresh(attempt)
    return attempt


def submit_quiz(db: Session, attempt_id: int, data: QuizAttemptSubmit) -> LearningQuizAttempt:
    attempt = db.query(LearningQuizAttempt).filter(LearningQuizAttempt.id == attempt_id).first()
    if not attempt:
        raise NotFoundException("LearningQuizAttempt", attempt_id)

    if attempt.status == "completed":
        raise BadRequestException("Quiz attempt has already been submitted.")

    assessment = get_assessment_by_id(db, attempt.assessment_id)
    questions = get_questions(db, attempt.assessment_id)

    try:
        user_answers = json.loads(data.answers)
    except (json.JSONDecodeError, TypeError):
        raise BadRequestException("Invalid answers format. Expected a JSON array.")

    if not isinstance(user_answers, list):
        raise BadRequestException("Invalid answers format. Expected a JSON array.")

    answer_map = {}
    for entry in user_answers:
        qid = entry.get("question_id")
        ans = entry.get("answer")
        if qid is not None:
            answer_map[qid] = ans

    total_points = 0
    earned_points = 0

    for question in questions:
        total_points += question.points
        user_answer = answer_map.get(question.id, "")
        if user_answer is not None and question.correct_answer is not None:
            if str(user_answer).strip().lower() == str(question.correct_answer).strip().lower():
                earned_points += question.points

    score = 0
    if total_points > 0:
        score = round((earned_points / total_points) * 100)

    passed = score >= assessment.passing_score

    attempt.answers = data.answers
    attempt.score = score
    attempt.passed = passed
    attempt.completed_at = datetime.utcnow()
    attempt.status = "completed"

    db.commit()
    db.refresh(attempt)
    return attempt


def get_quiz_attempts(
    db: Session,
    assessment_id: Optional[int] = None,
    employee_id: Optional[int] = None,
) -> list[LearningQuizAttempt]:
    query = db.query(LearningQuizAttempt)
    if assessment_id:
        query = query.filter(LearningQuizAttempt.assessment_id == assessment_id)
    if employee_id:
        query = query.filter(LearningQuizAttempt.employee_id == employee_id)
    return query.order_by(LearningQuizAttempt.started_at.desc()).all()


def get_quiz_attempt_by_id(db: Session, attempt_id: int) -> LearningQuizAttempt:
    attempt = db.query(LearningQuizAttempt).filter(LearningQuizAttempt.id == attempt_id).first()
    if not attempt:
        raise NotFoundException("LearningQuizAttempt", attempt_id)
    return attempt


def create_training_program(db: Session, data: TrainingProgramCreate, created_by: int, organization_id: Optional[int] = None) -> LearningTrainingProgram:
    program = LearningTrainingProgram(**data.model_dump(), created_by=created_by)
    if organization_id is not None:
        program.organization_id = organization_id
    db.add(program)
    db.commit()
    db.refresh(program)
    return program


def get_training_programs(
    db: Session,
    page: int = 1,
    per_page: int = 20,
    status: Optional[str] = None,
    organization_id: Optional[int] = None,
) -> dict:
    per_page = min(per_page, 100)
    try:
        query = db.query(LearningTrainingProgram)

        if status:
            query = query.filter(LearningTrainingProgram.status == status)

        if organization_id is not None:
            query = query.filter(LearningTrainingProgram.organization_id == organization_id)

        total = query.count()
        programs = query.order_by(LearningTrainingProgram.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

        items = []
        for p in programs:
            items.append({
                "id": p.id,
                "name": p.name,
                "description": p.description,
                "instructor_id": p.instructor_id,
                "start_date": p.start_date,
                "end_date": p.end_date,
                "status": p.status,
                "max_participants": p.max_participants,
                "participants_count": len(p.assignments) if p.assignments else 0,
                "created_by": p.created_by,
                "created_at": p.created_at,
                "updated_at": p.updated_at,
            })

        return {
            "total": total,
            "page": page,
            "per_page": per_page,
            "items": items,
        }
    except SQLAlchemyError:
        return {"total": 0, "page": page, "per_page": per_page, "items": []}


def get_training_program_by_id(db: Session, program_id: int, organization_id: Optional[int] = None) -> dict:
    program = db.query(LearningTrainingProgram).filter(LearningTrainingProgram.id == program_id).first()
    if not program:
        raise NotFoundException("LearningTrainingProgram", program_id)
    return {
        "id": program.id,
        "name": program.name,
        "description": program.description,
        "instructor_id": program.instructor_id,
        "start_date": program.start_date,
        "end_date": program.end_date,
        "status": program.status,
        "max_participants": program.max_participants,
        "participants_count": len(program.assignments) if program.assignments else 0,
        "created_by": program.created_by,
        "created_at": program.created_at,
        "updated_at": program.updated_at,
    }


def update_training_program(db: Session, program_id: int, data: TrainingProgramUpdate, organization_id: Optional[int] = None) -> LearningTrainingProgram:
    program = get_training_program_by_id(db, program_id, organization_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(program, field, value)
    db.commit()
    db.refresh(program)
    return program


def delete_training_program(db: Session, program_id: int, organization_id: Optional[int] = None) -> None:
    program = get_training_program_by_id(db, program_id, organization_id)
    db.delete(program)
    db.commit()


def assign_program(db: Session, data: ProgramAssignmentCreate, organization_id: Optional[int] = None) -> LearningTrainingProgramAssignment:
    get_training_program_by_id(db, data.program_id, organization_id)
    assignment = LearningTrainingProgramAssignment(**data.model_dump())
    if organization_id is not None:
        assignment.organization_id = organization_id
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    return assignment


def update_program_assignment(db: Session, assignment_id: int, data: ProgramAssignmentUpdate, organization_id: Optional[int] = None) -> LearningTrainingProgramAssignment:
    assignment = db.query(LearningTrainingProgramAssignment).filter(LearningTrainingProgramAssignment.id == assignment_id).first()
    if not assignment:
        raise NotFoundException("LearningTrainingProgramAssignment", assignment_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(assignment, field, value)
    db.commit()
    db.refresh(assignment)
    return assignment


def remove_program_assignment(db: Session, assignment_id: int, organization_id: Optional[int] = None) -> None:
    assignment = db.query(LearningTrainingProgramAssignment).filter(LearningTrainingProgramAssignment.id == assignment_id).first()
    if not assignment:
        raise NotFoundException("LearningTrainingProgramAssignment", assignment_id)
    db.delete(assignment)
    db.commit()


def get_program_assignments(db: Session, program_id: int, organization_id: Optional[int] = None) -> list[dict]:
    get_training_program_by_id(db, program_id, organization_id)
    assignments = db.query(LearningTrainingProgramAssignment).filter(
        LearningTrainingProgramAssignment.program_id == program_id
    ).all()
    items = []
    for a in assignments:
        items.append({
            "id": a.id,
            "program_id": a.program_id,
            "employee_id": a.employee_id,
            "status": a.status,
            "attended_at": a.attended_at,
            "created_at": a.created_at,
            "employee_name": a.employee.full_name if a.employee else None,
        })
    return items


def create_calendar_event(db: Session, data: CalendarEventCreate, created_by: int, organization_id: Optional[int] = None) -> LearningCalendarEvent:
    event = LearningCalendarEvent(**data.model_dump(), created_by=created_by)
    if organization_id is not None:
        event.organization_id = organization_id
    db.add(event)
    db.commit()
    db.refresh(event)
    return event


def get_calendar_events(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    organization_id: Optional[int] = None,
) -> list[LearningCalendarEvent]:
    query = db.query(LearningCalendarEvent)
    if start_date:
        query = query.filter(LearningCalendarEvent.event_date >= start_date)
    if end_date:
        query = query.filter(LearningCalendarEvent.event_date <= end_date)
    if organization_id is not None:
        query = query.filter(LearningCalendarEvent.organization_id == organization_id)
    return query.order_by(LearningCalendarEvent.event_date, LearningCalendarEvent.start_time).all()


def get_calendar_event_by_id(db: Session, event_id: int, organization_id: Optional[int] = None) -> LearningCalendarEvent:
    event = db.query(LearningCalendarEvent).filter(LearningCalendarEvent.id == event_id).first()
    if not event:
        raise NotFoundException("LearningCalendarEvent", event_id)
    return event


def update_calendar_event(db: Session, event_id: int, data: CalendarEventUpdate, organization_id: Optional[int] = None) -> LearningCalendarEvent:
    event = get_calendar_event_by_id(db, event_id, organization_id)
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(event, field, value)
    db.commit()
    db.refresh(event)
    return event


def delete_calendar_event(db: Session, event_id: int, organization_id: Optional[int] = None) -> None:
    event = get_calendar_event_by_id(db, event_id, organization_id)
    db.delete(event)
    db.commit()



def get_course_completion_report(db: Session) -> list[dict]:
    courses = db.query(LearningCourse).all()
    result = []
    for course in courses:
        total = db.query(LearningEnrollment).filter(LearningEnrollment.course_id == course.id).count()
        completed = db.query(LearningEnrollment).filter(
            LearningEnrollment.course_id == course.id,
            LearningEnrollment.status == "completed",
        ).count()
        rate = round((completed / total * 100) if total > 0 else 0.0, 2)
        result.append({
            "course_id": course.id,
            "course_name": course.course_name,
            "total_enrollments": total,
            "completed": completed,
            "completion_rate": rate,
        })
    return result


def get_skill_gap_analysis(db: Session) -> list[dict]:
    from app.modules.hr.models import Employee, Department
    from sqlalchemy import func
    skills = (
        db.query(
            Department.id.label("department_id"),
            Department.name.label("department_name"),
            LearningSkill.skill_name,
            func.avg(LearningSkill.proficiency_level).label("avg_level"),
            func.count(LearningSkill.id).label("employee_count"),
        )
        .join(Employee, LearningSkill.employee_id == Employee.id)
        .outerjoin(Department, Employee.department_id == Department.id)
        .group_by(Department.id, Department.name, LearningSkill.skill_name)
        .all()
    )
    dept_map = {}
    for s in skills:
        dept_id = s.department_id or 0
        dept_name = s.department_name or "Unassigned"
        if dept_id not in dept_map:
            dept_map[dept_id] = {
                "department_id": dept_id,
                "department_name": dept_name,
                "skills": []
            }
        gap = max(0, 5 - (s.avg_level or 0))
        gap_level = "low" if gap <= 1 else "medium" if gap <= 2 else "high"
        dept_map[dept_id]["skills"].append({
            "skill_id": None,
            "skill_name": s.skill_name,
            "employee_count": s.employee_count,
            "gap": round(gap, 2),
            "gap_level": gap_level
        })
    return list(dept_map.values())


def get_learning_dashboard(db: Session, organization_id: Optional[int] = None) -> dict:
    org_filter = [LearningCourse.organization_id == organization_id] if organization_id else []
    total_courses = db.query(LearningCourse).filter(*org_filter).count()
    active_courses = db.query(LearningCourse).filter(LearningCourse.status == "active", *org_filter).count()

    enroll_org_filter = [LearningEnrollment.organization_id == organization_id] if organization_id else []
    total_enrollments = db.query(LearningEnrollment).filter(*enroll_org_filter).count()
    completed_enrollments = db.query(LearningEnrollment).filter(LearningEnrollment.status == "completed", *enroll_org_filter).count()
    completion_rate = round((completed_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0.0, 2)

    cert_org_filter = [LearningCertification.organization_id == organization_id] if organization_id else []
    total_certifications = db.query(LearningCertification).filter(*cert_org_filter).count()

    skill_org_filter = [LearningSkill.organization_id == organization_id] if organization_id else []
    total_skills = db.query(LearningSkill).filter(*skill_org_filter).count()
    avg_skill_level = round(float(db.query(func.avg(LearningSkill.proficiency_level)).filter(*skill_org_filter).scalar() or 0.0), 2)

    attempt_org_filter = [LearningQuizAttempt.organization_id == organization_id] if organization_id else []
    in_progress_attempts = db.query(LearningQuizAttempt).filter(
        LearningQuizAttempt.status == "in_progress", *attempt_org_filter
    ).count()

    today = date.today()
    event_org_filter = [LearningCalendarEvent.organization_id == organization_id] if organization_id else []
    upcoming_events = db.query(LearningCalendarEvent).filter(
        LearningCalendarEvent.event_date >= today, *event_org_filter
    ).count()

    enrollment_trend = (
        db.query(
            func.to_char(LearningEnrollment.enrolled_at, "YYYY-MM").label("month"),
            func.count(LearningEnrollment.id).label("count"),
        )
        .filter(*enroll_org_filter)
        .group_by("month")
        .order_by("month")
        .all()
    )

    category_distribution = (
        db.query(
            LearningCourse.category,
            func.count(LearningCourse.id).label("count"),
        )
        .filter(LearningCourse.category.isnot(None), *org_filter)
        .group_by(LearningCourse.category)
        .all()
    )

    recent_enrollments_raw = (
        db.query(LearningEnrollment)
        .filter(*enroll_org_filter)
        .order_by(LearningEnrollment.enrolled_at.desc())
        .limit(10)
        .all()
    )

    recent_enrollments = []
    for e in recent_enrollments_raw:
        recent_enrollments.append({
            "id": e.id,
            "course_name": e.course.course_name if e.course else None,
            "employee_name": e.employee.full_name if e.employee else None,
            "status": e.status,
            "enrolled_at": str(e.enrolled_at) if e.enrolled_at else None,
        })

    return {
        "total_courses": total_courses,
        "active_courses": active_courses,
        "total_enrollments": total_enrollments,
        "completed_enrollments": completed_enrollments,
        "completion_rate": completion_rate,
        "total_certifications": total_certifications,
        "total_skills": total_skills,
        "avg_skill_level": avg_skill_level,
        "pending_assessments": in_progress_attempts,
        "upcoming_events": upcoming_events,
        "enrollment_trend": [{"month": m, "count": c} for m, c in enrollment_trend],
        "category_distribution": [{"category": cat, "count": cnt} for cat, cnt in category_distribution],
        "recent_enrollments": recent_enrollments,
    }


def export_course_completion_csv(db: Session) -> str:
    import io, csv
    data = get_course_completion_report(db)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Course ID", "Course Name", "Total Enrollments", "Completed Enrollments", "Completion Rate"])
    for row in data:
        writer.writerow([
            row["course_id"],
            row["course_name"],
            row["total_enrollments"],
            row["completed"],
            f"{row['completion_rate']}%"
        ])
    return output.getvalue()


def export_course_completion_excel(db: Session) -> bytes:
    import io
    from openpyxl import Workbook
    data = get_course_completion_report(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Completion"
    ws.append(["Course ID", "Course Name", "Total Enrollments", "Completed Enrollments", "Completion Rate"])
    for row in data:
        ws.append([
            row["course_id"],
            row["course_name"],
            row["total_enrollments"],
            row["completed"],
            f"{row['completion_rate']}%"
        ])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_certifications_csv(db: Session) -> str:
    import io, csv
    certs = get_certifications(db)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Employee ID", "Certification Name", "Issuing Organization", "Issue Date", "Expiry Date", "Credential URL", "Status"])
    for c in certs:
        writer.writerow([
            c.id,
            c.employee_id,
            c.certification_name,
            c.issuing_organization or "",
            c.issue_date,
            c.expiry_date or "",
            c.credential_url or "",
            c.status
        ])
    return output.getvalue()


def export_certifications_excel(db: Session) -> bytes:
    import io
    from openpyxl import Workbook
    certs = get_certifications(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Certifications"
    ws.append(["ID", "Employee ID", "Certification Name", "Issuing Organization", "Issue Date", "Expiry Date", "Credential URL", "Status"])
    for c in certs:
        ws.append([
            c.id,
            c.employee_id,
            c.certification_name,
            c.issuing_organization or "",
            str(c.issue_date),
            str(c.expiry_date) if c.expiry_date else "",
            c.credential_url or "",
            c.status
        ])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_skill_gap_csv(db: Session) -> str:
    import io, csv
    gaps = get_skill_gap_analysis(db)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Department Name", "Skill Name", "Employee Count", "Gap Level"])
    for dept in gaps:
        for s in dept["skills"]:
            writer.writerow([
                dept["department_name"],
                s["skill_name"],
                s["employee_count"],
                s["gap_level"]
            ])
    return output.getvalue()


def export_skill_gap_excel(db: Session) -> bytes:
    import io
    from openpyxl import Workbook
    gaps = get_skill_gap_analysis(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Skill Gap Analysis"
    ws.append(["Department Name", "Skill Name", "Employee Count", "Gap Level"])
    for dept in gaps:
        for s in dept["skills"]:
            ws.append([
                dept["department_name"],
                s["skill_name"],
                s["employee_count"],
                s["gap_level"]
            ])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
