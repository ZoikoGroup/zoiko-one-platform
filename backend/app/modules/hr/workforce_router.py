import csv
import io
from typing import Optional

from fastapi import APIRouter, Depends, Query, status
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session

from app.database import get_db
from app.core.dependencies import get_current_user, get_current_admin
from app.modules.hr import workforce_service
from app.modules.hr.schemas import (
    WfPlanCreate, WfPlanUpdate, WfPlanResponse,
    WfHeadcountCreate, WfHeadcountUpdate, WfHeadcountResponse,
    WfSuccessionCreate, WfSuccessionUpdate, WfSuccessionResponse,
    WfReportCreate, WfReportResponse,
    WfDashboardResponse,
    PaginatedWfPlans, PaginatedWfHeadcount, PaginatedWfSuccession, PaginatedWfReports,
)

workforce_router = APIRouter(prefix="/hr/workforce", tags=["Workforce Planning"])


# ════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ════════════════════════════════════════════════════════════════════════════

@workforce_router.get("/dashboard", response_model=WfDashboardResponse)
def workforce_dashboard(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    return workforce_service.get_dashboard(db, current_user.organization_id)


# ════════════════════════════════════════════════════════════════════════════
# PLANS CRUD
# ════════════════════════════════════════════════════════════════════════════

@workforce_router.get("/plans", response_model=PaginatedWfPlans)
def list_wf_plans(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=10000),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    department_id: Optional[int] = Query(None),
    plan_year: Optional[int] = Query(None),
    sort_by: str = Query("created_at"),
    sort_order: str = Query("desc"),
):
    return workforce_service.list_plans(
        db, current_user.organization_id, page, per_page,
        search, status, department_id, plan_year, sort_by, sort_order,
    )


@workforce_router.get("/plans/{plan_id}", response_model=WfPlanResponse)
def get_wf_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    return workforce_service.get_plan(db, current_user.organization_id, plan_id)


@workforce_router.post("/plans", response_model=WfPlanResponse, status_code=status.HTTP_201_CREATED)
def create_wf_plan(
    data: WfPlanCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    return workforce_service.create_plan(db, current_user.organization_id, data, current_user.id)


@workforce_router.put("/plans/{plan_id}", response_model=WfPlanResponse)
def update_wf_plan(
    plan_id: int,
    data: WfPlanUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    return workforce_service.update_plan(db, current_user.organization_id, plan_id, data, current_user.id)


@workforce_router.delete("/plans/{plan_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_wf_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    workforce_service.delete_plan(db, current_user.organization_id, plan_id)
    return Response(status_code=204)


# ════════════════════════════════════════════════════════════════════════════
# HEADCOUNT CRUD
# ════════════════════════════════════════════════════════════════════════════

@workforce_router.get("/headcount", response_model=PaginatedWfHeadcount)
def list_wf_headcounts(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=10000),
    department_id: Optional[int] = Query(None),
    fiscal_year: Optional[int] = Query(None),
    sort_by: str = Query("fiscal_year"),
    sort_order: str = Query("desc"),
):
    return workforce_service.list_headcounts(
        db, current_user.organization_id, page, per_page,
        department_id, fiscal_year, sort_by, sort_order,
    )


@workforce_router.get("/headcount/{hc_id}", response_model=WfHeadcountResponse)
def get_wf_headcount(
    hc_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    return workforce_service.get_headcount(db, current_user.organization_id, hc_id)


@workforce_router.post("/headcount", response_model=WfHeadcountResponse, status_code=status.HTTP_201_CREATED)
def create_wf_headcount(
    data: WfHeadcountCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    return workforce_service.create_headcount(db, current_user.organization_id, data, current_user.id)


@workforce_router.put("/headcount/{hc_id}", response_model=WfHeadcountResponse)
def update_wf_headcount(
    hc_id: int,
    data: WfHeadcountUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    return workforce_service.update_headcount(db, current_user.organization_id, hc_id, data, current_user.id)


@workforce_router.delete("/headcount/{hc_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_wf_headcount(
    hc_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    workforce_service.delete_headcount(db, current_user.organization_id, hc_id)
    return Response(status_code=204)


# ════════════════════════════════════════════════════════════════════════════
# SUCCESSION CRUD
# ════════════════════════════════════════════════════════════════════════════

@workforce_router.get("/succession", response_model=PaginatedWfSuccession)
def list_wf_successions(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=10000),
    readiness_level: Optional[str] = Query(None),
    risk_level: Optional[str] = Query(None),
    employee_id: Optional[int] = Query(None),
    sort_by: str = Query("created_at"),
    sort_order: str = Query("desc"),
):
    return workforce_service.list_successions(
        db, current_user.organization_id, page, per_page,
        readiness_level, risk_level, employee_id, sort_by, sort_order,
    )


@workforce_router.get("/succession/{succ_id}", response_model=WfSuccessionResponse)
def get_wf_succession(
    succ_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    return workforce_service.get_succession(db, current_user.organization_id, succ_id)


@workforce_router.post("/succession", response_model=WfSuccessionResponse, status_code=status.HTTP_201_CREATED)
def create_wf_succession(
    data: WfSuccessionCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    return workforce_service.create_succession(db, current_user.organization_id, data, current_user.id)


@workforce_router.put("/succession/{succ_id}", response_model=WfSuccessionResponse)
def update_wf_succession(
    succ_id: int,
    data: WfSuccessionUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    return workforce_service.update_succession(db, current_user.organization_id, succ_id, data, current_user.id)


@workforce_router.delete("/succession/{succ_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_wf_succession(
    succ_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    workforce_service.delete_succession(db, current_user.organization_id, succ_id)
    return Response(status_code=204)


# ════════════════════════════════════════════════════════════════════════════
# REPORTS
# ════════════════════════════════════════════════════════════════════════════

@workforce_router.get("/reports", response_model=PaginatedWfReports)
def list_wf_reports(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=10000),
    report_type: Optional[str] = Query(None),
    sort_by: str = Query("generated_at"),
    sort_order: str = Query("desc"),
):
    return workforce_service.list_reports(
        db, current_user.organization_id, page, per_page,
        report_type, sort_by, sort_order,
    )


@workforce_router.post("/reports/generate", response_model=WfReportResponse, status_code=status.HTTP_201_CREATED)
def generate_wf_report(
    data: WfReportCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin),
):
    return workforce_service.create_report(db, current_user.organization_id, data, current_user.id)


@workforce_router.get("/reports/export/csv")
def export_wf_csv(
    report_type: str = Query("workforce_summary"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rows = workforce_service.generate_report_data(db, current_user.organization_id, report_type)
    import csv, io
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=workforce_{report_type}.csv"},
    )


@workforce_router.get("/reports/export/excel")
def export_wf_excel(
    report_type: str = Query("workforce_summary"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rows = workforce_service.generate_report_data(db, current_user.organization_id, report_type)
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type
    if rows:
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[header])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=workforce_{report_type}.xlsx"},
    )


@workforce_router.get("/reports/export/pdf")
def export_wf_pdf(
    report_type: str = Query("workforce_summary"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rows = workforce_service.generate_report_data(db, current_user.organization_id, report_type)
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(503, detail="PDF generation requires reportlab. Install: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f"Workforce Report: {report_type}", styles["Title"]), Spacer(1, 12)]
    if rows:
        headers = list(rows[0].keys())
        table_data = [[Paragraph(h, styles["Normal"]) for h in headers]]
        for row in rows:
            table_data.append([Paragraph(str(row.get(h, "")), styles["Normal"]) for h in headers])
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=workforce_{report_type}.pdf"},
    )
